import streamlit as st
import pandas as pd
import pymysql
import math
import requests
import json
import base64
from io import BytesIO
from streamlit_drawable_canvas import st_canvas

st.set_page_config(page_title="E Report Board", layout="wide", page_icon="📋")

# ── Toast notification (after rerun) ─────────────────────────────────────
if "toast_msg" in st.session_state:
    st.toast(st.session_state["toast_msg"], icon=st.session_state.get("toast_icon", "ℹ️"))
    del st.session_state["toast_msg"]
    if "toast_icon" in st.session_state:
        del st.session_state["toast_icon"]

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Noto+Sans+TC:wght@400;700&family=Noto+Sans+SC:wght@400;700&display=swap');
html, body, [class*="css"], table {
    font-family: 'Noto Sans SC', 'Noto Sans TC', sans-serif !important;
}

/* Prevent date picker from auto-opening on re-render */
[data-baseweb="datepicker"] input:focus {
    outline: none !important;
}
[data-testid="stDialog"] [data-baseweb="datepicker"] input {
    pointer-events: auto;
}
/* ── Blue-White Theme ── */
.block-container { padding: 3.5rem 2rem 2rem 2rem !important; background: #F8FAFF; }
[data-testid="stSidebar"] {
    background: linear-gradient(180deg, #1565C0 0%, #1976D2 100%) !important;
}
[data-testid="stSidebar"] * { color: #fff !important; }
[data-testid="stSidebar"] [data-testid="stSelectbox"] > div > div {
    background: rgba(255,255,255,0.15) !important;
    border: 1px solid rgba(255,255,255,0.3) !important;
    color: #fff !important;
}
[data-testid="stSidebar"] button {
    background: rgba(255,255,255,0.9) !important;
    border: 1px solid rgba(255,255,255,0.4) !important;
    color: #1565C0 !important;
}
[data-testid="stSidebar"] button:hover {
    background: #fff !important;
    color: #0D47A1 !important;
}
[data-testid="stSidebar"] .stButton > button {
    background: rgba(255,255,255,0.9) !important;
    border: 1px solid rgba(255,255,255,0.4) !important;
    color: #1565C0 !important;
}
[data-testid="stSidebar"] .stButton > button:hover {
    background: #fff !important;
    color: #0D47A1 !important;
}
[data-testid="stSidebar"] .stButton > button p,
[data-testid="stSidebar"] .stButton > button span {
    color: #1565C0 !important;
}
[data-testid="stSidebar"] hr {
    border-color: rgba(255,255,255,0.2) !important;
}

/* Buttons */
.stButton > button {
    border-radius: 6px !important;
    font-size: 13px !important;
    border: 1px solid #1976D2 !important;
    color: #1976D2 !important;
    background: #fff !important;
    transition: all 0.2s;
}
.stButton > button:hover {
    background: #E3F2FD !important;
    border-color: #1565C0 !important;
}
.stButton > button[kind="primary"] {
    background: #1976D2 !important;
    color: #fff !important;
    border: none !important;
}
.stButton > button[kind="primary"]:hover {
    background: #1565C0 !important;
}

/* Tabs */
[data-testid="stTabs"] button[data-baseweb="tab"] {
    color: #1976D2 !important;
    font-weight: 600;
}
[data-testid="stTabs"] button[aria-selected="true"] {
    border-bottom-color: #1976D2 !important;
}

/* Alerts */
[data-testid="stAlert"] { border-radius: 8px; }

/* Compact grid table */
div[data-testid="stVerticalBlock"] > div.compact-row > div[data-testid="stVerticalBlock"] {
    gap: 0rem !important;
}
.compact-row [data-testid="stHorizontalBlock"] {
    border-bottom: 1px solid #E3F2FD;
    gap: 0 !important;
    background: #fff;
}
.compact-row [data-testid="stHorizontalBlock"]:hover {
    background: #F1F8FF;
}
.compact-row [data-testid="stHorizontalBlock"] > div[data-testid="stColumn"] {
    border-right: 1px solid #E3F2FD;
    padding: 6px 10px !important;
}
.compact-row [data-testid="stHorizontalBlock"] > div[data-testid="stColumn"]:last-child {
    border-right: none;
}
.compact-row p {
    font-size: 13px !important;
    margin: 0 !important;
    line-height: 1.4 !important;
    color: #333;
}

/* Selectbox & Input */
[data-testid="stSelectbox"] > div > div, .stTextInput > div > div > input {
    border-radius: 6px !important;
    border: 1px solid #BBDEFB !important;
}
[data-testid="stSelectbox"] > div > div:focus-within, .stTextInput > div > div > input:focus {
    border-color: #1976D2 !important;
    box-shadow: 0 0 0 1px #1976D2 !important;
}

/* Dialog scrollable content */
[data-testid="stDialog"] > div > div > div[data-testid="stVerticalBlock"] {
    max-height: 80vh !important;
    overflow-y: auto !important;
    padding-right: 4px !important;
}
/* Dialog detail table compact */
[data-testid="stDialog"] table td {
    font-size: 12px !important;
    padding: 4px 8px !important;
    word-break: break-word !important;
    max-width: 320px !important;
}
</style>
""", unsafe_allow_html=True)

# ── NocoDB API config ─────────────────────────────────────────────────────
NOCO_BASE = st.secrets["nocodb"]["base_url"]
NOCO_TOKEN = st.secrets["nocodb"]["token"]
NOCO_HEADERS = {"xc-token": NOCO_TOKEN}

# ── i18n ──────────────────────────────────────────────────────────────────
LANG = {
    "🇹🇭 ไทย": {
        "page_approve": "✍️ การเซ็นรับรอง",
        "page_report":  "📋 รายงานทั้งหมด",
        "filter":     "🗂️ กรองข้อมูล",
        "dept":       "แผนก",
        "all":        "ทั้งหมด",
        "search":     "🔍 ค้นหา",
        "search_ph":  "ชื่อรายงาน...",
        "rows":       "แถวต่อหน้า",
        "refresh":    "🔄 รีเฟรชข้อมูล",
        "auto":       "อัปเดตอัตโนมัติทุก 5 นาที",
        "no_data":    "ไม่พบข้อมูลในฐานข้อมูล",
        "no_result":  "ไม่พบรายการที่ค้นหา",
        "db_error":   "เชื่อมต่อฐานข้อมูลไม่ได้",
        "col_no":     "ลำดับ",
        "col_dept":   "แผนก",
        "col_name":   "ชื่อรายงาน",
        "col_form":   "📄 Form",
        "col_table":  "📊 Table",
        "col_date":   "อัปเดตล่าสุด",
        "open":       "🔗 เปิด",
        "page_info":  lambda p, tp, s, e, t: f"หน้า {p}/{tp}  •  แสดง {s}–{e} จาก {t}",
        "select_report": "เลือกรายงานที่ต้องการเซ็นรับรอง",
        "sign_title": "ลงลายเซ็น",
        "app_title":  "E Report Board — ระบบจัดการรายงานอิเล็กทรอนิกส์",
        "form_title": "กรอกแบบฟอร์ม",
        "form_submit": "📨 บันทึก",
        "form_success": "บันทึกข้อมูลสำเร็จ",
        "form_error": "ไม่สามารถบันทึกข้อมูลได้",
        "form_loading": "กำลังโหลด schema...",
        "col_comment":  "คำแนะนำ",
        "comment_title": "📝 คำแนะนำสำหรับแบบฟอร์มนี้",
        "comment_save":  "💾 บันทึกคำแนะนำ",
        "comment_saved": "บันทึกคำแนะนำเรียบร้อยแล้ว",
        "comment_ph":    "กรอกคำแนะนำหรือวิธีการกรอกฟอร์มนี้...",
        "sign_clear": "ล้างลายเซ็น",
        "sign_submit": "✅ ยืนยันเซ็นรับรอง",
        "sign_success": "บันทึกลายเซ็นเรียบร้อยแล้ว",
        "sign_error":  "ไม่สามารถบันทึกลายเซ็นได้",
        "sign_empty":  "กรุณาเซ็นลายเซ็นก่อนยืนยัน",
        "records":     "รายการข้อมูล",
        "no_records":  "ไม่พบรายการในรายงานนี้",
        "signer_name": "ชื่อผู้เซ็น",
        "tbl_no":      "No.",
        "tbl_date":    "วันที่สร้าง",
        "tbl_report":  "ชื่อ Report",
        "tbl_recorder":"记录者 Recorder",
        "tbl_sig_status": "สถานะลายเซ็น",
        "tbl_sign":    "เซ็น",
        "existing_sig": "📋 ลายเซ็นที่บันทึกแล้ว",
        "signed":      "✅ เซ็นแล้ว",
        "not_signed":  "⏳ ยังไม่ได้เซ็น",
        "no_sig_field":"ไม่พบ field ลายเซ็น",
        "sign_incomplete": "⚠️ กรุณาเซ็นให้ครบทุกช่อง: ",
        "sign_confirm_success": "✅ เซ็นรับรองสำเร็จทุกช่อง",
        "sign_confirm_fail": "❌ เซ็นรับรองไม่สำเร็จบางช่อง",
        "tbl_delete":  "แก้ไข",
        "delete_success": "บันทึกการแก้ไขสำเร็จ",
        "delete_fail": "ไม่สามารถบันทึกการแก้ไขได้",
        "edit_title":  "แก้ไขรายการ",
        "edit_save":   "💾 บันทึก",
        "tbl_col_edit": "แก้ไข",
        "tbl_col_del":  "ลบ",
        "del_confirm":  "ยืนยันลบรายการนี้?",
        "del_success":  "ลบรายการเรียบร้อยแล้ว",
        "del_fail":     "ไม่สามารถลบรายการได้",
        "tab1_date_filter": "เลือกวันที่ดูรายการ",
        "tab1_loading": "กำลังโหลดข้อมูลจากทุกตาราง...",
        "tab1_no_records": "ไม่พบรายการในวันที่เลือก",
        "tab1_col_report": "ชื่อรายงาน",
        "tab1_col_dept": "แผนก",
                "cfg_none_field": "(ไม่กำหนด)",
        "cfg_qr_title":   "📷 QR Code Auto-fill",
        "cfg_qr_caption": "กำหนดฟิลด์ที่ต้องการให้รับค่าจากการสแกน QR Code ด้วยกล้อง",
        "cfg_qr_delimiter": "Delimiter (ตัวคั่นข้อมูลใน QR)",
        "cfg_qr_col_field": "ฟิลด์",
        "cfg_qr_col_segment": "Segment (0=ทั้งหมด, 1=ส่วนแรก, 2=ส่วนที่สอง...)",
        "cfg_qr_add_field": "➕ เพิ่มฟิลด์ QR",
        "cfg_qr_remove": "ลบ",
        "print_btn":   "🖨️",
        "print_title": "🖨️ พิมพ์รายการ",
        "print_do":    "🖨️ พิมพ์",
        "print_locked":"เซ็นไม่ครบ ไม่สามารถพิมพ์ได้",
        "tbl_print":   "พิมพ์",
        "admin_login":  "เข้าสู่ Admin",
        "admin_logout": "ออกจาก Admin",
        "admin_pw_ph":  "🔒 รหัสผ่าน Admin",
        "admin_wrong_pw": "รหัสผ่านไม่ถูกต้อง",
        "admin_mode":   "✅ Admin mode",
        "cfg_btn_help": "จัดการโครงสร้าง field",
        "cfg_dialog_caption": "กำหนดว่า field ไหนต้องแสดง, บังคับกรอก, และลำดับการแสดงผล",
        "cfg_col_field":  "Field",
        "cfg_col_visible":"แสดง",
        "cfg_col_req":    "บังคับกรอก",
        "cfg_col_order":  "ลำดับ",
        "cfg_col_ph":     "ลายน้ำ",
        "cfg_save":       "💾 บันทึก config",
        "cfg_saved":      "✅ บันทึก form config เรียบร้อย",
        "cfg_err_schema": "ดึง schema ไม่ได้",
        "col_cfg":        "⚙️ Config",
        "col_export":     "📥 Export",
        "export_btn_help": "ส่งออกข้อมูลทั้งตารางเป็น Excel",
        "export_generating": "กำลังดึงข้อมูลและสร้างไฟล์ Excel...",
        "export_download": "⬇️ ดาวน์โหลด Excel",
        "export_no_data": "ไม่พบข้อมูลในตารางนี้",
        "export_error":   "ไม่สามารถสร้างไฟล์ Excel ได้",
        "cfg_notify_title": "🔔 แจ้งเตือนการกรอกข้อมูล (Webhook)",
        "cfg_notify_enable": "เปิดใช้งานแจ้งเตือน",
        "cfg_notify_webhook": "Webhook URL",
        "cfg_notify_webhook_help": "ระบบจะส่ง HTTP POST เป็น JSON ไปยัง URL นี้",
        "cfg_notify_template": "ข้อความที่จะส่ง (ตอนยังไม่กรอก)",
        "cfg_notify_template_help": "ตัวแปรที่ใช้ได้: {report_name}, {start_hour}, {now}",
        "cfg_notify_success_enable": "แจ้งเตือนด้วยเมื่อมีการกรอกข้อมูลแล้ว",
        "cfg_notify_success_template": "ข้อความที่จะส่ง (ตอนกรอกแล้ว)",
        "cfg_notify_start_time": "เวลาเริ่มแจ้งเตือน",
        "cfg_notify_interval": "แจ้งเตือนซ้ำทุก (ชั่วโมง)",
        "cfg_notify_test": "📤 ทดสอบส่งข้อความ",
        "cfg_notify_test_ok": "✅ ส่งข้อความทดสอบสำเร็จ",
        "cfg_notify_test_fail": "❌ ส่งข้อความทดสอบไม่สำเร็จ: ",
        "required_missing": "⚠️ กรุณากรอกข้อมูลให้ครบ: ",
        "upload_label":     "อัปโหลดไฟล์/รูปภาพ",
        "upload_error":     "อัปโหลดไฟล์ไม่สำเร็จ",
        "dlg_comment":  "📝 คำแนะนำแบบฟอร์ม",
        "dlg_form":     "📋 กรอกแบบฟอร์ม",
        "dlg_sign":     "📝 เซ็นรับรอง",
        "dlg_edit":     "✏️ แก้ไขรายการ",
        "dlg_print":    "🖨️ พิมพ์รายการ",
        "dlg_delete":   "🗑️ ยืนยันการลบ",
        "dlg_cfg":      "⚙️ จัดการโครงสร้างแบบฟอร์ม",
        "cfg_sec_date_title": "กำหนด Date Auto-fill",
        "cfg_primary_date":   "Field วันที่หลัก",
        "cfg_secondary_dates":"Field วันที่รอง",
        "cfg_offset_months":  "ระยะเวลา (เดือน)",
        "cfg_none":           "(ไม่กำหนด)",
        "cfg_ph_hint":        "ลายน้ำของ field นี้",
        "cfg_autofill_enable": "เปิด Auto-fill",
        "cfg_date_format_title": "กำหนด Date Format",
        "cfg_date_format_label": "รูปแบบวันที่",
        "cfg_range_min": "ค่าต่ำสุด",
        "cfg_range_max": "ค่าสูงสุด",
        "cfg_range_title": "Standard Range",
        "range_error": "⚠️ ค่าไม่อยู่ใน range ที่กำหนด",
        "cfg_allow_edit_title": "🔓 อนุญาตให้แก้ไขหลังส่งแบบฟอร์ม",
        "cfg_allow_edit_label": "อนุญาตให้ user ทั่วไปแก้ไขข้อมูลที่บันทึกแล้วได้",
        "cfg_allow_edit_caption": "เมื่อเปิด option นี้ จะมีปุ่ม ✏️ ให้ user ธรรมดาแก้ไขได้โดยไม่ต้อง login admin",
    },
    "🇬🇧 English": {
        "page_approve": "✍️ Approval Signing",
        "page_report":  "📋 All Reports",
        "filter":     "🗂️ Filter",
        "dept":       "Department",
        "all":        "All",
        "search":     "🔍 Search",
        "search_ph":  "Report name...",
        "rows":       "Rows per page",
        "refresh":    "🔄 Refresh",
        "auto":       "Auto-refresh every 5 minutes",
        "no_data":    "No data found",
        "no_result":  "No results found",
        "db_error":   "Cannot connect to database",
        "col_no":     "#",
        "col_dept":   "Department",
        "col_name":   "Report Name",
        "col_form":   "📄 Form",
        "col_table":  "📊 Table",
        "col_date":   "Last Updated",
        "open":       "🔗 Open",
        "page_info":  lambda p, tp, s, e, t: f"Page {p}/{tp}  •  Showing {s}–{e} of {t}",
        "select_report": "Select report to sign",
        "sign_title": "Signature",
        "app_title":  "E Report Board — Electronic Report Management",
        "form_title": "Fill Form",
        "form_submit": "📨 Submit",
        "form_success": "Record saved successfully",
        "form_error": "Failed to save record",
        "form_loading": "Loading schema...",
        "col_comment":  "Guide",
        "comment_title": "📝 Form Instructions",
        "comment_save":  "💾 Save Instructions",
        "comment_saved": "Instructions saved",
        "comment_ph":    "Enter instructions or guidance for filling this form...",
        "sign_clear": "Clear signature",
        "sign_submit": "✅ Confirm Signing",
        "sign_success": "Signature saved successfully",
        "sign_error":  "Failed to save signature",
        "sign_empty":  "Please sign before confirming",
        "records":     "Records",
        "no_records":  "No records found for this report",
        "signer_name": "Signer name",
        "tbl_no":      "No.",
        "tbl_date":    "Created Date",
        "tbl_report":  "Report Name",
        "tbl_recorder":"Recorder",
        "tbl_sig_status": "Signature Status",
        "tbl_sign":    "Sign",
        "existing_sig": "📋 Existing Signatures",
        "signed":      "✅ Signed",
        "not_signed":  "⏳ Not signed",
        "no_sig_field":"No signature field found",
        "sign_incomplete": "⚠️ Please sign all fields: ",
        "sign_confirm_success": "✅ All signatures submitted successfully",
        "sign_confirm_fail": "❌ Some signatures failed to save",
        "tbl_delete":  "Edit",
        "delete_success": "Changes saved successfully",
        "delete_fail": "Failed to save changes",
        "edit_title":  "Edit Record",
        "edit_save":   "💾 Save",
        "tbl_col_edit": "Edit Record",
        "tbl_col_del":  "Delete Record",
        "del_confirm":  "Confirm delete this record?",
        "del_success":  "Record deleted successfully",
        "del_fail":     "Failed to delete record",
        "tab1_date_filter": "Select date to view records",
        "tab1_loading": "Loading records from all tables...",
        "tab1_no_records": "No records found for selected date",
        "tab1_col_report": "Report Name",
        "tab1_col_dept": "Department",
        "cfg_field_as_workflow": "Field → Part Number (WORKFLOW)",
        "cfg_field_as_lot": "Field → Lot Name (cascade)",
        "cfg_none_field": "(none)",
        "cfg_qr_title":   "📷 QR Code Auto-fill",
        "cfg_qr_caption": "Configure fields to receive values from QR code camera scan",
        "cfg_qr_delimiter": "Delimiter (separator in QR data)",
        "cfg_qr_col_field": "Field",
        "cfg_qr_col_segment": "Segment (0=full, 1=first, 2=second...)",
        "cfg_qr_add_field": "➕ Add QR field",
        "cfg_qr_remove": "Remove",
        "print_btn":   "🖨️",
        "print_title": "🖨️ Print Record",
        "print_do":    "🖨️ Print",
        "print_locked":"Signatures incomplete — cannot print",
        "tbl_print":   "Print",
        "admin_login":  "Enter Admin",
        "admin_logout": "Exit Admin",
        "admin_pw_ph":  "🔒 Admin Password",
        "admin_wrong_pw": "Incorrect password",
        "admin_mode":   "✅ Admin mode",
        "cfg_btn_help": "Manage field config",
        "cfg_dialog_caption": "Configure which fields to show, mark as required, and set display order",
        "cfg_col_field":  "Field",
        "cfg_col_visible":"Visible",
        "cfg_col_req":    "Required",
        "cfg_col_order":  "Order",
        "cfg_col_ph":     "Placeholder",
        "cfg_save":       "💾 Save config",
        "cfg_saved":      "✅ Form config saved",
        "cfg_err_schema": "Failed to load schema",
        "col_cfg":        "⚙️ Config",
        "col_export":     "📥 Export",
        "export_btn_help": "Export full table data to Excel",
        "export_generating": "Fetching data and building Excel file...",
        "export_download": "⬇️ Download Excel",
        "export_no_data": "No data found in this table",
        "export_error":   "Failed to build Excel file",
        "cfg_notify_title": "🔔 Data Entry Reminder (Webhook)",
        "cfg_notify_enable": "Enable reminder",
        "cfg_notify_webhook": "Webhook URL",
        "cfg_notify_webhook_help": "Sends an HTTP POST with JSON to this URL",
        "cfg_notify_template": "Message template (when not filled)",
        "cfg_notify_template_help": "Available variables: {report_name}, {start_hour}, {now}",
        "cfg_notify_success_enable": "Also notify when data has been filled",
        "cfg_notify_success_template": "Message template (when filled)",
        "cfg_notify_start_time": "Reminder start time",
        "cfg_notify_interval": "Repeat every (hours)",
        "cfg_notify_test": "📤 Send test message",
        "cfg_notify_test_ok": "✅ Test message sent",
        "cfg_notify_test_fail": "❌ Failed to send test message: ",
        "required_missing": "⚠️ Please fill in required fields: ",
        "upload_label":     "Upload file/image",
        "upload_error":     "File upload failed",
        "dlg_comment":  "📝 Form Instructions",
        "dlg_form":     "📋 Fill Form",
        "dlg_sign":     "📝 Sign Approval",
        "dlg_edit":     "✏️ Edit Record",
        "dlg_print":    "🖨️ Print Record",
        "dlg_delete":   "🗑️ Confirm Delete",
        "dlg_cfg":      "⚙️ Form Field Config",
        "cfg_sec_date_title": "Date Auto-fill Config",
        "cfg_primary_date":   "Primary date field",
        "cfg_secondary_dates":"Secondary date fields",
        "cfg_offset_months":  "Offset (months)",
        "cfg_none":           "(none)",
        "cfg_ph_hint":        "Placeholder text for field",
        "cfg_autofill_enable": "Enable Auto-fill",
        "cfg_date_format_title": "Date Format",
        "cfg_date_format_label": "Date format",
        "cfg_range_min": "Min value",
        "cfg_range_max": "Max value",
        "cfg_range_title": "Standard Range",
        "range_error": "⚠️ Value out of allowed range",
        "cfg_allow_edit_title": "🔓 Allow editing after submission",
        "cfg_allow_edit_label": "Allow regular users to edit submitted records",
        "cfg_allow_edit_caption": "When enabled, an ✏️ edit button appears for all users (no admin login required)",
    },
    "🇨🇳 中文": {
        "page_approve": "✍️ 签名认证",
        "page_report":  "📋 所有报表",
        "filter":     "🗂️ 筛选",
        "dept":       "部门",
        "all":        "全部",
        "search":     "🔍 搜索",
        "search_ph":  "报表名称...",
        "rows":       "每页行数",
        "refresh":    "🔄 刷新数据",
        "auto":       "每5分钟自动更新",
        "no_data":    "数据库中未找到数据",
        "no_result":  "未找到相关记录",
        "db_error":   "无法连接数据库",
        "col_no":     "序号",
        "col_dept":   "部门",
        "col_name":   "报表名称",
        "col_form":   "📄 表单",
        "col_table":  "📊 报表",
        "col_date":   "最后更新",
        "open":       "🔗 打开",
        "page_info":  lambda p, tp, s, e, t: f"第 {p}/{tp} 页  •  显示 {s}–{e} 共 {t} 条",
        "select_report": "选择需要签名的报表",
        "sign_title": "签名",
        "app_title":  "E Report Board — 电子报表管理系统",
        "form_title": "填写表单",
        "form_submit": "📨 提交",
        "form_success": "记录保存成功",
        "form_error": "保存失败",
        "form_loading": "加载中...",
        "col_comment":  "说明",
        "comment_title": "📝 表单说明",
        "comment_save":  "💾 保存说明",
        "comment_saved": "说明已保存",
        "comment_ph":    "输入填写此表单的说明或注意事项...",
        "sign_clear": "清除签名",
        "sign_submit": "✅ 确认签名",
        "sign_success": "签名保存成功",
        "sign_error":  "无法保存签名",
        "sign_empty":  "请先签名再确认",
        "records":     "记录列表",
        "no_records":  "该报表无记录",
        "signer_name": "签名人姓名",
        "tbl_no":      "序号",
        "tbl_date":    "创建日期",
        "tbl_report":  "报表名称",
        "tbl_recorder":"记录者",
        "tbl_sig_status": "签名状态",
        "tbl_sign":    "签名",
        "existing_sig": "📋 已有签名",
        "signed":      "✅ 已签",
        "not_signed":  "⏳ 未签",
        "no_sig_field":"未找到签名字段",
        "sign_incomplete": "⚠️ 请完成所有签名字段: ",
        "sign_confirm_success": "✅ 所有签名提交成功",
        "sign_confirm_fail": "❌ 部分签名保存失败",
        "tbl_delete":  "编辑",
        "delete_success": "保存成功",
        "delete_fail": "保存失败",
        "edit_title":  "编辑记录",
        "edit_save":   "💾 保存",
        "tbl_col_edit": "编辑记录",
        "tbl_col_del":  "删除记录",
        "del_confirm":  "确认删除此记录？",
        "del_success":  "记录已删除",
        "del_fail":     "删除失败",
        "tab1_date_filter": "选择日期查看记录",
        "tab1_loading": "正在加载所有表格数据...",
        "tab1_no_records": "所选日期无记录",
        "tab1_col_report": "报表名称",
        "tab1_col_dept": "部门",
        "cfg_field_as_workflow": "字段 → 料号 (WORKFLOW)",
        "cfg_field_as_lot": "字段 → 批号 (cascade)",
        "cfg_none_field": "(不设置)",
        "cfg_qr_title":   "📷 QR Code 自动填充",
        "cfg_qr_caption": "设置哪些字段通过摄像头扫描 QR Code 来自动填充",
        "cfg_qr_delimiter": "Delimiter（QR 数据中的分隔符）",
        "cfg_qr_col_field": "字段",
        "cfg_qr_col_segment": "Segment（0=完整, 1=第一段, 2=第二段...）",
        "cfg_qr_add_field": "➕ 添加字段",
        "cfg_qr_remove": "删除",
        "print_btn":   "🖨️",
        "print_title": "🖨️ 打印记录",
        "print_do":    "🖨️ 打印",
        "print_locked":"签名未完成，无法打印",
        "tbl_print":   "打印",
        "admin_login":  "进入管理员",
        "admin_logout": "退出管理员",
        "admin_pw_ph":  "🔒 管理员密码",
        "admin_wrong_pw": "密码错误",
        "admin_mode":   "✅ 管理员模式",
        "cfg_btn_help": "管理字段配置",
        "cfg_dialog_caption": "设置字段的显示、必填及排列顺序",
        "cfg_col_field":  "字段",
        "cfg_col_visible":"显示",
        "cfg_col_req":    "必填",
        "cfg_col_order":  "顺序",
        "cfg_col_ph":     "水印",
        "cfg_save":       "💾 保存配置",
        "cfg_saved":      "✅ 表单配置已保存",
        "cfg_err_schema": "无法加载字段结构",
        "col_cfg":        "⚙️ 配置",
        "col_export":     "📥 导出",
        "export_btn_help": "将整个表格数据导出为 Excel",
        "export_generating": "正在获取数据并生成 Excel 文件...",
        "export_download": "⬇️ 下载 Excel",
        "export_no_data": "该表格没有数据",
        "export_error":   "生成 Excel 文件失败",
        "cfg_notify_title": "🔔 填表提醒 (Webhook)",
        "cfg_notify_enable": "启用提醒",
        "cfg_notify_webhook": "Webhook URL",
        "cfg_notify_webhook_help": "系统会向此 URL 发送 HTTP POST JSON",
        "cfg_notify_template": "发送的消息内容（未填写时）",
        "cfg_notify_template_help": "可用变量: {report_name}, {start_hour}, {now}",
        "cfg_notify_success_enable": "填写完成后也发送通知",
        "cfg_notify_success_template": "发送的消息内容（已填写时）",
        "cfg_notify_start_time": "提醒开始时间",
        "cfg_notify_interval": "重复提醒间隔（小时）",
        "cfg_notify_test": "📤 发送测试消息",
        "cfg_notify_test_ok": "✅ 测试消息已发送",
        "cfg_notify_test_fail": "❌ 测试消息发送失败: ",
        "required_missing": "⚠️ 请填写必填项: ",
        "upload_label":     "上传文件/图片",
        "upload_error":     "文件上传失败",
        "dlg_comment":  "📝 表单说明",
        "dlg_form":     "📋 填写表单",
        "dlg_sign":     "📝 签名认证",
        "dlg_edit":     "✏️ 编辑记录",
        "dlg_print":    "🖨️ 打印记录",
        "dlg_delete":   "🗑️ 确认删除",
        "dlg_cfg":      "⚙️ 表单字段配置",
        "cfg_sec_date_title": "日期自动填充配置",
        "cfg_primary_date":   "主日期字段",
        "cfg_secondary_dates":"从属日期字段",
        "cfg_offset_months":  "间隔月数",
        "cfg_none":           "(不设置)",
        "cfg_ph_hint":        "该字段的水印文字",
        "cfg_autofill_enable": "启用自动填充",
        "cfg_date_format_title": "日期格式",
        "cfg_date_format_label": "日期格式",
        "cfg_range_min": "最小值",
        "cfg_range_max": "最大值",
        "cfg_range_title": "标准范围",
        "range_error": "⚠️ 数值超出允许范围",
        "cfg_allow_edit_title": "🔓 允许提交后编辑",
        "cfg_allow_edit_label": "允许普通用户编辑已提交的记录",
        "cfg_allow_edit_caption": "启用后，所有用户都可看到 ✏️ 编辑按钮，无需管理员登录",
    },
}

# ── DB ────────────────────────────────────────────────────────────────────
CATALOG_TABLE = "ereport_catalog"

def get_db_config():
    cfg = st.secrets["mysql"]
    return dict(host=cfg["host"], port=int(cfg["port"]),
                database=cfg["database"], user=cfg["user"], password=cfg["password"],
                connect_timeout=5, read_timeout=10)

def ensure_catalog_table(conn):
    """Create the new-system report catalog when it does not exist."""
    with conn.cursor() as cursor:
        cursor.execute(f"""
            CREATE TABLE IF NOT EXISTS `{CATALOG_TABLE}` (
                id INT NOT NULL AUTO_INCREMENT PRIMARY KEY,
                dept VARCHAR(255) NOT NULL,
                report_name VARCHAR(500) NOT NULL,
                table_id VARCHAR(100) NOT NULL UNIQUE,
                from_link VARCHAR(1000) DEFAULT NULL,
                table_link VARCHAR(1000) DEFAULT NULL,
                updated_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP
                    ON UPDATE CURRENT_TIMESTAMP,
                nc_order INT NOT NULL DEFAULT 0
            ) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci
        """)

@st.cache_data(ttl=300)
def load_data():
    try:
        conn = pymysql.connect(**get_db_config())
        ensure_catalog_table(conn)
        conn.commit()
        df = pd.read_sql("""
            SELECT
                dept,
                report_name AS name,
                from_link AS From_link,
                table_link AS Table_link,
                table_id AS Table_ID,
                DATE(updated_at)   AS updated
            FROM `ereport_catalog`
            ORDER BY nc_order
        """, conn)
        conn.close()
        return df, None
    except Exception as e:
        return pd.DataFrame(), str(e)


# ── NocoDB API helpers ────────────────────────────────────────────────────
def noco_get_records(table_id, limit=100, offset=0):
    """Fetch records from a NocoDB table by Table_ID."""
    url = f"{NOCO_BASE}/api/v2/tables/{table_id}/records?offset={offset}&limit={limit}"
    try:
        resp = requests.get(url, headers=NOCO_HEADERS, timeout=10)
        if resp.status_code == 200:
            return resp.json(), None
        return None, f"HTTP {resp.status_code}: {resp.text}"
    except Exception as e:
        return None, str(e)

def noco_get_all_records(table_id, page_size=200):
    """Fetch ALL records from a NocoDB table via pagination."""
    all_recs = []
    offset = 0
    while True:
        data, err = noco_get_records(table_id, limit=page_size, offset=offset)
        if err or not data:
            break
        recs = data.get("list", [])
        all_recs.extend(recs)
        page_info = data.get("pageInfo", {})
        if page_info.get("isLastPage", True) or not recs:
            break
        offset += page_size
    return all_recs

def noco_get_records_by_date(table_id, date_str, limit=500):
    """Fetch all records from a table and filter by date (Python-side) due to NocoDB CreatedAt filter limitation."""
    url = f"{NOCO_BASE}/api/v2/tables/{table_id}/records?limit={limit}"
    try:
        resp = requests.get(url, headers=NOCO_HEADERS, timeout=15)
        if resp.status_code != 200:
            return [], f"HTTP {resp.status_code}"
        all_recs = resp.json().get("list", [])
        # Filter by date — convert UTC CreatedAt to UTC+7 and compare date
        import re as _re3
        from datetime import datetime as _dt2, timezone as _tz2, timedelta as _td2
        _local = _tz2(_td2(hours=7))
        filtered = []
        for r in all_recs:
            cat = str(r.get("CreatedAt", "") or "")
            if not cat:
                continue
            try:
                dstr = cat.strip().replace("Z", "+00:00")
                dstr = _re3.sub(r'(\.\d{6})\d+', r'\1', dstr)
                dt = _dt2.fromisoformat(dstr)
                if dt.tzinfo is None:
                    dt = dt.replace(tzinfo=_tz2.utc)
                dt_local = dt.astimezone(_local)
                if dt_local.strftime("%Y-%m-%d") == date_str:
                    filtered.append(r)
            except Exception:
                pass
        return filtered, None
    except Exception as e:
        return [], str(e)

def noco_upload_signature(table_id, record_id, field_name, signature_base64, filename=None):
    """Upload signature image to a specific field of a record via NocoDB API."""
    upload_url = f"{NOCO_BASE}/api/v2/storage/upload"
    img_data = base64.b64decode(signature_base64)
    fname = filename or f"{field_name}.png"
    files = {"file": (fname, BytesIO(img_data), "image/png")}
    try:
        upload_resp = requests.post(upload_url, headers=NOCO_HEADERS, files=files, timeout=15)
        if upload_resp.status_code != 200:
            return False, f"Upload failed: {upload_resp.status_code}"
        attachment = upload_resp.json()

        # Update the record with signature in the target field
        patch_url = f"{NOCO_BASE}/api/v2/tables/{table_id}/records"
        payload = {"Id": record_id, field_name: attachment}
        patch_resp = requests.patch(patch_url, headers=NOCO_HEADERS,
                                    json=payload, timeout=10)
        if patch_resp.status_code == 200:
            return True, None
        return False, f"Patch failed: {patch_resp.status_code}"
    except Exception as e:
        return False, str(e)

def get_signature_fields(records):
    """Find all field names that start with 'signature' (case-insensitive)."""
    if not records:
        return []
    first = records[0]
    return [k for k in first.keys() if k.lower().startswith("signature")]

def _download_image_bytes(url, retries=3):
    """Download image bytes with retries (NocoDB can be slow/flaky under many sequential requests)."""
    last_err = None
    for attempt in range(retries):
        try:
            resp = requests.get(url, headers=NOCO_HEADERS, timeout=20)
            if resp.status_code == 200 and resp.content:
                return resp.content
            last_err = f"HTTP {resp.status_code}"
        except Exception as e:
            last_err = str(e)
    return None

def build_excel_with_images(records, sheet_name="Data"):
    """Build an .xlsx file (in memory) from NocoDB records.
    Attachment/signature fields (list of dicts with url) are embedded as
    actual images in the cell; other list/dict fields are dumped as JSON text.
    Returns (bytes, error)."""
    import openpyxl
    import concurrent.futures
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.utils import get_column_letter

    if not records:
        return None, "no records"

    # Collect all field names across records, keep first-seen order
    fieldnames = []
    for r in records:
        for k in r.keys():
            if k not in fieldnames:
                fieldnames.append(k)

    def is_attachment_list(val):
        return isinstance(val, list) and len(val) > 0 and isinstance(val[0], dict) \
            and any(("url" in v or "path" in v or "signedUrl" in v or "signedPath" in v) for v in val)

    # ── Pass 1: figure out which cells need an image and resolve their URL ──
    img_jobs = []  # list of dicts: row_idx, col_idx, url, fallback_title
    cell_text = {}  # (row_idx, col_idx) -> text to write (non-image cells)

    for row_offset, rec in enumerate(records):
        row_idx = row_offset + 2
        for col_idx, fname in enumerate(fieldnames, start=1):
            val = rec.get(fname, "")
            if val is None:
                val = ""
            if is_attachment_list(val):
                att = val[0]
                mime = att.get("mimetype", "") or ""
                # NOTE: "path" is the permanent /download/... URL. "signedPath"/
                # "signedUrl" are short-lived signed URLs (NocoDB "dltemp") that
                # can expire within seconds — using them causes intermittent
                # missing images. Prefer "path" first.
                url = (att.get("path") or att.get("url") or
                       att.get("signedPath") or att.get("signedUrl") or "")
                title = att.get("title", att.get("filename", ""))
                if url and mime.startswith("image/"):
                    url = url.replace("\\", "/")
                    if not url.startswith("http"):
                        url = f"{NOCO_BASE}/{url.lstrip('/')}"
                    img_jobs.append({"row": row_idx, "col": col_idx, "url": url, "title": title})
                else:
                    titles = [a.get("title", a.get("filename", "")) for a in val]
                    cell_text[(row_idx, col_idx)] = ", ".join(t for t in titles if t)
            elif isinstance(val, (dict, list)):
                cell_text[(row_idx, col_idx)] = json.dumps(val, ensure_ascii=False)
            else:
                cell_text[(row_idx, col_idx)] = val

    # ── Pass 2: download all images in parallel (with retries) ──────────────
    img_bytes_map = {}
    if img_jobs:
        with concurrent.futures.ThreadPoolExecutor(max_workers=8) as ex:
            future_to_job = {ex.submit(_download_image_bytes, job["url"]): job for job in img_jobs}
            for fut in concurrent.futures.as_completed(future_to_job):
                job = future_to_job[fut]
                try:
                    data = fut.result()
                except Exception:
                    data = None
                img_bytes_map[(job["row"], job["col"])] = (data, job["title"])

    # ── Pass 3: build the workbook ────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name[:31] if sheet_name else "Data"

    for col_idx, fname in enumerate(fieldnames, start=1):
        ws.cell(row=1, column=col_idx, value=fname)

    IMG_ROW_HEIGHT = 90
    IMG_COL_WIDTH = 16
    rows_with_image = set()

    for (row_idx, col_idx), (data, title) in img_bytes_map.items():
        if data:
            try:
                img = XLImage(BytesIO(data))
                img.height = IMG_ROW_HEIGHT
                img.width = IMG_ROW_HEIGHT
                col_letter = get_column_letter(col_idx)
                img.anchor = f"{col_letter}{row_idx}"
                ws.add_image(img)
                rows_with_image.add(row_idx)
                continue
            except Exception:
                pass
        # download failed — fall back to filename text
        if title:
            ws.cell(row=row_idx, column=col_idx, value=title)

    for (row_idx, col_idx), text_val in cell_text.items():
        ws.cell(row=row_idx, column=col_idx, value=text_val)

    for row_idx in rows_with_image:
        ws.row_dimensions[row_idx].height = IMG_ROW_HEIGHT * 0.75

    for col_idx, fname in enumerate(fieldnames, start=1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = max(IMG_COL_WIDTH, min(40, len(str(fname)) + 4))

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue(), None

def noco_delete_record(table_id, record_id):
    """Delete a record from NocoDB table."""
    url = f"{NOCO_BASE}/api/v2/tables/{table_id}/records"
    try:
        resp = requests.delete(url, headers=NOCO_HEADERS, json={"Id": record_id}, timeout=10)
        if resp.status_code == 200:
            return True, None
        return False, f"HTTP {resp.status_code}: {resp.text}"
    except Exception as e:
        return False, str(e)

def noco_update_record(table_id, record_id, fields):
    """Update fields of a record in NocoDB."""
    url = f"{NOCO_BASE}/api/v2/tables/{table_id}/records"
    payload = {"Id": record_id, **fields}
    try:
        resp = requests.patch(url, headers=NOCO_HEADERS, json=payload, timeout=10)
        if resp.status_code == 200:
            return True, None
        return False, f"HTTP {resp.status_code}: {resp.text}"
    except Exception as e:
        return False, str(e)

COMMENTS_FILE = "comments.json"
FORM_CONFIG_FILE = "form_config.json"

def load_comments():
    """Load comments dict from JSON file."""
    try:
        with open(COMMENTS_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
        return {k: v for k, v in data.items() if not k.startswith("__")}
    except Exception:
        return {}

def save_comments(comments: dict):
    """Save comments dict to JSON file."""
    try:
        existing = {}
        try:
            with open(COMMENTS_FILE, "r", encoding="utf-8") as f:
                existing = json.load(f)
        except Exception:
            pass
        existing.update(comments)
        with open(COMMENTS_FILE, "w", encoding="utf-8") as f:
            json.dump(existing, f, ensure_ascii=False, indent=2)
        return True
    except Exception:
        return False

def load_form_config():
    """Load form field config dict from JSON file."""
    try:
        with open(FORM_CONFIG_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}

def decode_qr_from_image(image_bytes):
    """Decode QR code data from image bytes using cv2.
    Returns (decoded_text, error_string)."""
    try:
        import cv2
        import numpy as np
        from PIL import Image as _PILImage
        img_pil = _PILImage.open(BytesIO(image_bytes)).convert("RGB")
        img_np = np.array(img_pil)
        img_bgr = cv2.cvtColor(img_np, cv2.COLOR_RGB2BGR)
        detector = cv2.QRCodeDetector()
        data, _, _ = detector.detectAndDecode(img_bgr)
        if data:
            return data, None
        return None, "ไม่พบ QR Code ในภาพ"
    except Exception as e:
        return None, str(e)


def qr_webrtc_scanner(widget_key: str, result_key: str):
    """Show a camera input for QR scanning (browser-side, no server camera needed).
    User takes a photo; QR is decoded server-side with cv2.
    Returns decoded string if found, else None."""
    cam_img = st.camera_input("📷 ถ่ายหรืออัปโหลด QR Code",
                              key=f"cam_{widget_key}",
                              label_visibility="visible")
    if cam_img is not None:
        decoded, err = decode_qr_from_image(cam_img.getvalue())
        if err:
            st.warning(f"⚠️ {err}")
            return None
        st.session_state[result_key] = decoded
        return decoded
    return st.session_state.get(result_key)

def send_webhook_message(webhook_url, text):
    """Send a notification to a generic webhook endpoint as JSON."""
    if not webhook_url:
        return False, "no webhook url"
    payload = {"message": text, "text": text}
    try:
        current_url = webhook_url
        for _ in range(5):
            resp = requests.post(current_url, json=payload, timeout=10, allow_redirects=False)
            if resp.status_code in (301, 302, 303, 307, 308) and "Location" in resp.headers:
                current_url = resp.headers["Location"]
                continue
            break
        if 200 <= resp.status_code < 300:
            return True, None
        return False, f"HTTP {resp.status_code}: {resp.text[:200]}"
    except Exception as e:
        return False, str(e)

def save_form_config(config: dict):
    """Save form field config dict to JSON file."""
    try:
        existing = load_form_config()
        existing.update(config)
        with open(FORM_CONFIG_FILE, "w", encoding="utf-8") as f:
            json.dump(existing, f, ensure_ascii=False, indent=2)
        return True
    except Exception:
        return False


@st.cache_data(ttl=60)
def noco_get_fields(table_id):
    """Fetch field schema for a NocoDB table via /api/v1/db/meta/tables/{id}."""
    url = f"{NOCO_BASE}/api/v1/db/meta/tables/{table_id}"
    try:
        resp = requests.get(url, headers=NOCO_HEADERS, timeout=10)
        if resp.status_code == 200:
            data = resp.json()
            cols = data.get("columns", [])
            # Sort by NocoDB field order so new/reordered fields appear correctly
            cols.sort(key=lambda c: c.get("order") or 9999)
            return cols, None
        return [], f"HTTP {resp.status_code}"
    except Exception as e:
        return [], str(e)

def noco_create_table(base_id, title, table_name, columns):
    """Create a NocoDB table with columns through the metadata API."""
    if not base_id:
        return None, "กรุณากำหนด nocodb.base_id ใน secrets.toml"
    url = f"{NOCO_BASE}/api/v1/db/meta/projects/{base_id}/tables"
    payload = {"table_name": table_name,
               "title": title,
               "columns": columns}
    try:
        resp = requests.post(url, headers=NOCO_HEADERS, json=payload, timeout=30)
        if resp.status_code in (200, 201):
            data = resp.json()
            table_id = data.get("id") or data.get("table_id") or data.get("tableId")
            return table_id, None
        return None, f"HTTP {resp.status_code}: {resp.text}"
    except Exception as e:
        return None, str(e)

def register_catalog_report(dept, report_name, table_id):
    """Register a newly-created NocoDB table in the existing report catalog."""
    conn = None
    try:
        conn = pymysql.connect(**get_db_config())
        ensure_catalog_table(conn)
        with conn.cursor() as cursor:
            cursor.execute("""
                INSERT INTO `ereport_catalog`
                    (dept, report_name, table_id, updated_at, nc_order)
                VALUES (%s, %s, %s, NOW(),
                    (SELECT COALESCE(MAX(x.nc_order), 0) + 1 FROM
                        (SELECT nc_order FROM `ereport_catalog`) x))
            """, (dept, report_name, table_id))
        conn.commit()
        return True, None
    except Exception as e:
        if conn:
            conn.rollback()
        return False, str(e)
    finally:
        if conn:
            conn.close()

def noco_delete_table(table_id):
    """Delete a NocoDB table through the metadata API."""
    url = f"{NOCO_BASE}/api/v1/db/meta/tables/{table_id}"
    try:
        resp = requests.delete(url, headers=NOCO_HEADERS, timeout=30)
        if resp.status_code in (200, 204):
            return True, None
        return False, f"HTTP {resp.status_code}: {resp.text[:200]}"
    except Exception as e:
        return False, str(e)

def delete_catalog_report(table_id):
    """Remove a report from the new-system catalog."""
    conn = None
    try:
        conn = pymysql.connect(**get_db_config())
        ensure_catalog_table(conn)
        with conn.cursor() as cursor:
            cursor.execute(f"DELETE FROM `{CATALOG_TABLE}` WHERE table_id = %s", (table_id,))
        conn.commit()
        return True, None
    except Exception as e:
        if conn:
            conn.rollback()
        return False, str(e)
    finally:
        if conn:
            conn.close()

def noco_create_record(table_id, fields):
    """Create a new record in a NocoDB table."""
    url = f"{NOCO_BASE}/api/v2/tables/{table_id}/records"
    try:
        resp = requests.post(url, headers=NOCO_HEADERS, json=fields, timeout=10)
        if resp.status_code == 200:
            return True, None
        return False, f"HTTP {resp.status_code}: {resp.text}"
    except Exception as e:
        return False, str(e)


def noco_upload_attachment(uploaded_file):
    """Upload a file to NocoDB storage and return attachment object."""
    upload_url = f"{NOCO_BASE}/api/v2/storage/upload"
    try:
        files = {"file": (uploaded_file.name, uploaded_file.getvalue(), uploaded_file.type)}
        resp = requests.post(upload_url, headers=NOCO_HEADERS, files=files, timeout=30)
        if resp.status_code == 200:
            return resp.json(), None
        return None, f"HTTP {resp.status_code}: {resp.text}"
    except Exception as e:
        return None, str(e)

# ── Load ──────────────────────────────────────────────────────────────────
df, error = load_data()

# ── Resolve language early so T is available inside dialogs ──────────────
if "lang_key" not in st.session_state:
    st.session_state["lang_key"] = list(LANG.keys())[0]
T = LANG[st.session_state["lang_key"]]
# ── Sidebar (logo + language + nav) ──────────────────────────────────────
with st.sidebar:
    st.markdown("""
    <div style="display:flex;align-items:center;gap:10px;padding:8px 0 4px 0">
        <span style="font-size:28px">📋</span>
        <div>
            <div style="font-size:13px;font-weight:700;letter-spacing:1px;line-height:1.2">E Report Board</div>
            <div style="font-size:11px;opacity:0.6;letter-spacing:2px">MANAGER</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    st.divider()

    lang_key = st.selectbox("🌐 Language", list(LANG.keys()),
                             index=list(LANG.keys()).index(st.session_state["lang_key"]),
                             label_visibility="collapsed")
    st.session_state["lang_key"] = lang_key
    T = LANG[lang_key]

    st.divider()
    # ── Admin mode ────────────────────────────────────────────────────────
    if "is_admin" not in st.session_state:
        st.session_state["is_admin"] = False

    if not st.session_state["is_admin"]:
        admin_pw = st.text_input("🔒 Admin Password", type="password", key="admin_pw_input",
                                 label_visibility="collapsed", placeholder=T["admin_pw_ph"])
        if st.button(T["admin_login"], use_container_width=True, key="admin_login_btn"):
            correct_pw = st.secrets.get("admin", {}).get("password", "admin1234")
            if admin_pw == correct_pw:
                st.session_state["is_admin"] = True
                st.rerun()
            else:
                st.error(T["admin_wrong_pw"])
    else:
        st.success(T["admin_mode"])
        if st.button(T["admin_logout"], use_container_width=True, key="admin_logout_btn"):
            st.session_state["is_admin"] = False
            st.rerun()

    st.divider()
    if st.session_state.get("is_admin", False) and st.button(
            "➕ สร้างแบบฟอร์มใหม่", use_container_width=True, key="new_form_btn"):
        st.session_state.pop("pending_dialog", None)
        st.session_state["pending_form_builder"] = True
        st.rerun()

# ── Validation ────────────────────────────────────────────────────────────
if error:
    st.error(f"{T['db_error']}: {error}")
    st.stop()
if df.empty:
    st.warning(T["no_data"])

# ── Top Ribbon ────────────────────────────────────────────────────────────
st.markdown(f"""
<div style="background:linear-gradient(135deg, #1565C0, #1976D2, #42A5F5);
            padding:18px 28px;border-radius:12px;margin-bottom:18px;
            display:flex;align-items:center;gap:16px;
            box-shadow:0 4px 12px rgba(21,101,192,0.25)">
    <span style="font-size:36px">📋</span>
    <div>
        <div style="font-size:22px;font-weight:700;color:#fff;letter-spacing:0.5px">E Report Board</div>
        <div style="font-size:12px;color:rgba(255,255,255,0.85);margin-top:2px">{T['app_title']}</div>
    </div>
</div>
""", unsafe_allow_html=True)

# ── Page Tabs ─────────────────────────────────────────────────────────────
tab1, tab2 = st.tabs([T["page_approve"], T["page_report"]])


# ══════════════════════════════════════════════════════════════════════════
# TAB 1: การเซ็นรับรอง (Approval Signing)
# ══════════════════════════════════════════════════════════════════════════

@st.dialog(" ", width="large")
def comment_dialog():
    """Modal dialog for editing form instructions."""
    cdata = st.session_state.get("comment_record")
    if not cdata:
        return
    table_id = cdata["table_id"]
    report_name = cdata["report_name"]
    st.subheader(T["dlg_comment"])
    st.markdown(f"**{report_name}**")

    comments = load_comments()
    current = comments.get(table_id, "")
    new_comment = st.text_area(T["comment_ph"], value=current,
                               height=200, key=f"comment_input_{table_id}")

    if st.button(T["comment_save"], type="primary", use_container_width=True):
        ok = save_comments({table_id: new_comment})
        if ok:
            del st.session_state["comment_record"]
            st.session_state["toast_msg"] = T["comment_saved"]
            st.session_state["toast_icon"] = "💬"
            st.rerun()


@st.dialog(" ", width="large")
def form_dialog():
    """Modal dialog for creating a new record via NocoDB API."""
    form_data = st.session_state.get("form_record")
    if not form_data:
        return
    table_id = form_data["table_id"]
    report_name = form_data["report_name"]
    st.subheader(T["dlg_form"])
    st.markdown(f"**{report_name}**")

    # Show comment/instructions if exists
    comments = load_comments()
    comment_text = comments.get(table_id, "")
    if comment_text:
        st.info(f"📋 {comment_text}")

    # Fetch field schema
    fields, err = noco_get_fields(table_id)
    if not fields:
        # Fallback: infer from existing records
        rec_data, rec_err = noco_get_records(table_id, limit=1)
        if not rec_err and rec_data and rec_data.get("list"):
            sample = rec_data["list"][0]
            SKIP = {"id", "nc_order", "createdat", "updatedat", "created_at", "updated_at"}
            fields = [
                {"title": k, "uidt": "SingleLineText", "required": False}
                for k in sample.keys()
                if k.lower() not in SKIP and not k.lower().startswith("signature")
            ]
        else:
            # No records and no meta — render a generic free-form entry
            st.warning("ไม่สามารถโหลด schema ได้ กรุณากรอกข้อมูลในรูปแบบ JSON")
            raw = st.text_area("JSON payload", value="{}", key="form_raw_json")
            if st.button(T["form_submit"], type="primary", use_container_width=True, key="form_submit_raw"):
                try:
                    payload = json.loads(raw)
                    success, e = noco_create_record(table_id, payload)
                    if success:
                        del st.session_state["form_record"]
                        st.session_state["toast_msg"] = T["form_success"]
                        st.session_state["toast_icon"] = "✅"
                        st.rerun()
                    else:
                        st.error(f"{T['form_error']}: {e}")
                except json.JSONDecodeError as je:
                    st.error(f"JSON invalid: {je}")
            return

    import re
    from datetime import datetime, date, time

    # Skip system/signature/readonly fields (Attachment is handled separately)
    SKIP_UIDT = {"ID", "CreatedTime", "LastModifiedTime", "CreatedBy", "LastModifiedBy", "Order", "Rollup", "Lookup", "Formula"}
    editable_fields = [
        f for f in fields
        if f.get("title", "").lower() not in ("id", "nc_order")
        and not f.get("title", "").lower().startswith("signature")
        and not f.get("system", False)
        and f.get("uidt") not in SKIP_UIDT
    ]

    # ── Apply Admin form config (override required, visible, order) ───────
    form_cfg = load_form_config().get(table_id, {})
    field_cfg = {fc["title"]: fc for fc in form_cfg.get("fields", [])}
    date_autofill = form_cfg.get("date_autofill", {})
    primary_date_field = date_autofill.get("primary", "")
    sec_by_field = {s["field"]: s for s in date_autofill.get("secondaries", []) if s.get("enabled")}
    date_fmt = form_cfg.get("date_format", "YYYY-MM-DD")
    # QR autofill config — build a lookup: field_title -> segment index
    qr_cfg = form_cfg.get("qr_autofill", {})
    qr_delimiter = qr_cfg.get("delimiter", "|")
    qr_field_map = {m["field"]: m["segment"] for m in qr_cfg.get("field_mappings", []) if m.get("field")}
    editable_fields = [
        f for f in editable_fields
        if field_cfg.get(f.get("title", ""), {}).get("visible", True)
    ]
    editable_fields.sort(key=lambda f: field_cfg.get(f.get("title", ""), {}).get("order", f.get("order") or 9999))

    from datetime import date as _date, time, timedelta
    from dateutil.relativedelta import relativedelta

    # ── Pre-compute auto-fill: inject secondary session_state before render
    if primary_date_field and sec_by_field:
        primary_val_now = st.session_state.get(f"form_{primary_date_field}")
        prev_key = f"_af_prev_{table_id}"
        if primary_val_now != st.session_state.get(prev_key):
            st.session_state[prev_key] = primary_val_now
            for sec_fname, sec_cfg in sec_by_field.items():
                offset_months = sec_cfg.get("offset_months", 1)
                if primary_val_now and isinstance(primary_val_now, _date):
                    st.session_state[f"form_{sec_fname}"] = (
                        primary_val_now + relativedelta(months=offset_months)
                    )
                else:
                    st.session_state.pop(f"form_{sec_fname}", None)

    values = {}
    required_fields = []
    primary_field_obj = None

    for field in editable_fields:
        fname = field.get("field_name", field.get("title", ""))
        ftype = field.get("uidt", "SingleLineText")
        cfg = field_cfg.get(fname, {})
        required = cfg.get("required", bool(field.get("rqd", 0)))
        placeholder = cfg.get("placeholder", "")
        date_fmt = cfg.get("date_format", "YYYY-MM-DD") or "YYYY-MM-DD"
        DATE_INPUT_FMT_MAP = {
            "YYYY-MM-DD": "YYYY-MM-DD", "DD/MM/YYYY": "DD/MM/YYYY",
            "MM/DD/YYYY": "MM/DD/YYYY", "DD-MM-YYYY": "DD/MM/YYYY",
            "YYYY/MM/DD": "YYYY/MM/DD",
            "YYYY-MM (วันที่=1)": "YYYY-MM-DD",
            "YYYY (วันที่=01-01)": "YYYY-MM-DD",
        }
        date_input_fmt = DATE_INPUT_FMT_MAP.get(date_fmt, "YYYY-MM-DD")
        if required:
            required_fields.append(fname)
        label = fname + (" *" if required else "")
        fname_lower = fname.lower()

        if ftype == "SingleLineText":
            date_keywords = ["date", "日期", "วันที่"]
            time_keywords = ["time", "时间", "เวลา"]
            if any(fname_lower == w or fname_lower.startswith(w + " ") or fname_lower.endswith(" " + w)
                   for w in date_keywords):
                ftype = "Date"
            elif any(fname_lower == w or fname_lower.startswith(w + " ") or fname_lower.endswith(" " + w)
                     for w in time_keywords):
                ftype = "Time"

        if ftype in ("Date", "DateTime"):
            values[fname] = st.date_input(label, format=date_input_fmt, key=f"form_{fname}")
        elif ftype in ("Time",):
            from st_mui.time_picker import time_picker as _tp
            from datetime import datetime as _now_dt, time as _time
            _now_time = _now_dt.now().time().replace(second=0, microsecond=0)
            values[fname] = _tp(label=label, value=_now_time, ampm=False, key=f"form_{fname}")
        elif ftype in ("Number", "Decimal", "Currency", "Percent"):
            step = 1 if ftype == "Number" else 0.01
            # Build range hint from config
            rmin_op  = cfg.get("range_min_op", "—")
            rmin_val = str(cfg.get("range_min_val", "") or "").strip()
            rmax_op  = cfg.get("range_max_op", "—")
            rmax_val = str(cfg.get("range_max_val", "") or "").strip()
            range_parts = []
            if rmin_op != "—" and rmin_val:
                range_parts.append(f"{rmin_op} {rmin_val}")
            if rmax_op != "—" and rmax_val:
                range_parts.append(f"{rmax_op} {rmax_val}")
            range_hint = "  ,  ".join(range_parts) if range_parts else None

            values[fname] = st.number_input(
                label,
                value=None,
                step=step,
                format="%d" if ftype == "Number" else "%g",
                help=range_hint,
                key=f"form_{fname}"
            )
            # Real-time range warning
            cur_val = values[fname]
            if cur_val is not None and range_parts:
                warn_msgs = []
                try:
                    if rmin_op != "—" and rmin_val:
                        t = float(rmin_val)
                        ok = (rmin_op == ">=" and cur_val >= t) or (rmin_op == ">" and cur_val > t) or (rmin_op == "=" and cur_val == t)
                        if not ok:
                            warn_msgs.append(f"ต้อง {rmin_op} {rmin_val}")
                    if rmax_op != "—" and rmax_val:
                        t = float(rmax_val)
                        ok = (rmax_op == "<=" and cur_val <= t) or (rmax_op == "<" and cur_val < t) or (rmax_op == "=" and cur_val == t)
                        if not ok:
                            warn_msgs.append(f"ต้อง {rmax_op} {rmax_val}")
                except (ValueError, TypeError):
                    pass
                if warn_msgs:
                    st.warning(f"⚠️ {fname}: " + "  ,  ".join(warn_msgs))
        elif ftype in ("Checkbox",):
            values[fname] = st.checkbox(label, key=f"form_{fname}")
        elif ftype in ("Attachment",):
            uploaded = st.file_uploader(
                label, type=["jpg","jpeg","png","gif","pdf","xlsx","docx","csv"],
                accept_multiple_files=True,
                key=f"form_{fname}", help=T["upload_label"]
            )
            values[fname] = uploaded or []  # handled specially at submit
        elif ftype in ("LongText",):
            values[fname] = st.text_area(label, placeholder=placeholder or None, key=f"form_{fname}")
        elif ftype in ("SingleSelect",):
            raw_opts = field.get("colOptions", {})
            opts_raw = raw_opts.get("options", "") if isinstance(raw_opts, dict) else ""
            if isinstance(opts_raw, list):
                opts = [o.get("title", str(o)) for o in opts_raw]
            else:
                import re as _re
                opts = [o.strip().strip("'") for o in _re.split(r",(?=')", str(opts_raw)) if o.strip()]
            values[fname] = st.selectbox(label, [""] + opts, key=f"form_{fname}")
        elif ftype in ("MultiSelect",):
            raw_opts = field.get("colOptions", {})
            opts_raw = raw_opts.get("options", "") if isinstance(raw_opts, dict) else ""
            if isinstance(opts_raw, list):
                opts = [o.get("title", str(o)) for o in opts_raw]
            else:
                import re as _re
                opts = [o.strip().strip("'") for o in _re.split(r",(?=')", str(opts_raw)) if o.strip()]
            sel = st.multiselect(label, opts, key=f"form_{fname}")
            values[fname] = ",".join(sel)
        else:
            # ── QR camera button for QR-enabled fields ──────────────────────
            if fname in qr_field_map:
                qr_show_key = f"qr_show_{table_id}_{fname}"
                qr_result_key = f"qr_result_{table_id}_{fname}"
                txt_col, cam_col = st.columns([9, 1])
                with txt_col:
                    values[fname] = st.text_input(label, placeholder=placeholder or None,
                                                  key=f"form_{fname}")
                with cam_col:
                    st.markdown("<div style='padding-top:28px'>", unsafe_allow_html=True)
                    if st.button("📷", key=f"qr_btn_{table_id}_{fname}", help="สแกน QR Code"):
                        st.session_state[qr_show_key] = not st.session_state.get(qr_show_key, False)
                    st.markdown("</div>", unsafe_allow_html=True)
                if st.session_state.get(qr_show_key, False):
                    st.caption("📷 ส่องกล้องไปที่ QR Code — จะอ่านอัตโนมัติ")
                    decoded = qr_webrtc_scanner(
                        widget_key=f"qr_webrtc_{table_id}_{fname}",
                        result_key=qr_result_key,
                    )
                    if decoded:
                        seg_idx = qr_field_map.get(fname, 0)
                        parts = decoded.split(qr_delimiter) if qr_delimiter and seg_idx > 0 else [decoded]
                        fill_val = decoded if seg_idx == 0 else (parts[seg_idx - 1].strip() if seg_idx <= len(parts) else decoded)
                        st.session_state[f"form_{fname}"] = fill_val
                        st.session_state[qr_show_key] = False
                        st.rerun()
            else:
                values[fname] = st.text_input(label, placeholder=placeholder or None,
                                              key=f"form_{fname}")
    st.markdown("")
    if st.button(T["form_submit"], type="primary", use_container_width=True, key="form_submit_btn"):

        # Validate required fields
        missing = [f for f in required_fields if not values.get(f)]
        if missing:
            st.error(T["required_missing"] + ", ".join(missing))
            return

        # Validate number ranges
        range_errors = []
        for fname, v in values.items():
            fcfg = field_cfg.get(fname, {})
            rmin_op  = fcfg.get("range_min_op", "—")
            rmin_val = str(fcfg.get("range_min_val", "") or "").strip()
            rmax_op  = fcfg.get("range_max_op", "—")
            rmax_val = str(fcfg.get("range_max_val", "") or "").strip()
            if v is None or v == "":
                continue
            try:
                num_v = float(v)
                for op, val_str, label in [(rmin_op, rmin_val, T["cfg_range_min"]),
                                           (rmax_op, rmax_val, T["cfg_range_max"])]:
                    if op == "—" or not val_str:
                        continue
                    threshold = float(val_str)
                    ok = (
                        (op == ">=" and num_v >= threshold) or
                        (op == ">"  and num_v >  threshold) or
                        (op == "<=" and num_v <= threshold) or
                        (op == "<"  and num_v <  threshold) or
                        (op == "="  and num_v == threshold)
                    )
                    if not ok:
                        range_errors.append(f"{fname}: {v} ไม่ตรงเงื่อนไข {op}{val_str}")
            except (ValueError, TypeError):
                pass
        if range_errors:
            st.error(T["range_error"] + "\n- " + "\n- ".join(range_errors))
            return

        # Filter out empty values and convert date/time to string
        payload = {}
        upload_errors = []
        for k, v in values.items():
            if v is None or v == "" or v == 0.0:
                continue
            if isinstance(v, _date):
                fcfg = field_cfg.get(k, {})
                fmt = fcfg.get("date_format", "YYYY-MM-DD") or "YYYY-MM-DD"
                if fmt == "YYYY-MM (วันที่=1)":
                    payload[k] = v.strftime("%Y-%m-01")
                elif fmt == "YYYY (วันที่=01-01)":
                    payload[k] = v.strftime("%Y-01-01")
                elif fmt == "DD-MM-YYYY":
                    payload[k] = v.strftime("%d-%m-%Y")
                else:
                    payload[k] = str(v)
            elif isinstance(v, time):
                payload[k] = str(v)
            elif hasattr(v, "read") or hasattr(v, "getvalue"):
                # Single file upload
                attachment, up_err = noco_upload_attachment(v)
                if up_err:
                    upload_errors.append(f"{k}: {up_err}")
                elif attachment:
                    payload[k] = [attachment] if isinstance(attachment, dict) else attachment
            elif isinstance(v, list) and v and hasattr(v[0], "getvalue"):
                # Multiple file uploads
                att_list = []
                for uf in v:
                    attachment, up_err = noco_upload_attachment(uf)
                    if up_err:
                        upload_errors.append(f"{k}: {up_err}")
                    elif attachment:
                        att_list.append(attachment if isinstance(attachment, dict) else attachment[0])
                if att_list:
                    payload[k] = att_list
            else:
                payload[k] = v

        if upload_errors:
            st.error(f"{T['upload_error']}: {', '.join(upload_errors)}")
            return

        success, err = noco_create_record(table_id, payload)
        if success:
            del st.session_state["form_record"]
            st.session_state["toast_msg"] = T["form_success"]
            st.session_state["toast_icon"] = "✅"
            st.rerun()
        else:
            st.error(f"{T['form_error']}: {err}")


@st.dialog(" ", width="large")
def sign_dialog():
    """Modal dialog for signing a record."""
    sign_data = st.session_state.get("sign_record")
    if not sign_data:
        st.warning("No record selected")
        return
    st.subheader(T["dlg_sign"])

    rec = sign_data["record"]
    rec_id = sign_data["record_id"]
    tbl_id = sign_data["table_id"]
    s_fields = sign_data["sig_fields"]
    rec_idx = sign_data["idx"]

    st.markdown(f"**Record #{rec_idx+1}**")

    # Record detail as styled HTML table
    detail_keys = [k for k in rec.keys()
                   if not k.lower().startswith("signature")
                   and k.lower() not in ("nc_order", "createdat", "updatedat")]

    detail_rows = ""
    attachment_fields = {}  # fname → list of attachment objects
    for i, k in enumerate(detail_keys):
        val = rec.get(k, "") or ""
        if not val:
            continue
        # Detect attachment (list of dicts with url/signedUrl)
        if isinstance(val, list) and len(val) > 0 and isinstance(val[0], dict):
            attachment_fields[k] = val
            continue
        bg = "rgba(0,0,0,0.02)" if i % 2 == 0 else "transparent"
        k_short = (k[:30] + "…") if len(k) > 30 else k
        detail_rows += f"""<tr style="background:{bg}">
            <td style="padding:6px 10px;font-weight:600;max-width:160px;width:160px;
                       overflow:hidden;text-overflow:ellipsis;white-space:nowrap;
                       border-right:1px solid rgba(128,128,128,0.15);color:#555"
                title="{k}">{k_short}</td>
            <td style="padding:6px 10px;max-width:260px;overflow:hidden;
                       text-overflow:ellipsis;white-space:nowrap;word-break:break-word"
                title="{val}">{val}</td>
        </tr>"""

    st.markdown(f"""
    <table style="width:100%;border-collapse:collapse;font-size:12px;
                  border:1px solid rgba(128,128,128,0.2);border-radius:8px;overflow:hidden;margin-bottom:10px;table-layout:fixed">
        <tbody>{detail_rows}</tbody>
    </table>
    """, unsafe_allow_html=True)

    # Show attachment fields as images/links
    if attachment_fields:
        for att_key, att_list in attachment_fields.items():
            st.markdown(f"**{att_key}**")
            thumbs_html = '<div style="display:flex;flex-wrap:wrap;gap:8px;align-items:flex-start;margin-bottom:8px">'
            for att in att_list:
                url = att.get("signedUrl", att.get("url", att.get("path", "")))
                if not url:
                    continue
                if not url.startswith("http"):
                    url = f"{NOCO_BASE}/{url.lstrip('/')}"
                mime = att.get("mimetype", "")
                title = att.get("title", att.get("filename", "file"))
                if mime.startswith("image/"):
                    thumbs_html += f'''<a href="{url}" target="_blank" title="{title}">
                        <img src="{url}" style="height:140px;width:auto;border:2px solid #BBDEFB;
                             border-radius:8px;object-fit:cover;cursor:zoom-in;transition:all 0.2s"
                             onmouseover="this.style.borderColor='#1976D2';this.style.boxShadow='0 4px 14px rgba(25,118,210,0.35)';this.style.transform='scale(1.03)'"
                             onmouseout="this.style.borderColor='#BBDEFB';this.style.boxShadow='none';this.style.transform='scale(1)'"
                        /></a>'''
                else:
                    thumbs_html += f'<a href="{url}" target="_blank" style="display:inline-flex;align-items:center;gap:4px;padding:6px 12px;border:1px solid #BBDEFB;border-radius:6px;font-size:13px;text-decoration:none;color:#1976D2">📎 {title}</a>'
            thumbs_html += '</div>'
            st.markdown(thumbs_html, unsafe_allow_html=True)

    st.divider()

    # Show existing signatures if already signed
    has_existing_sig = False
    for sf in s_fields:
        val = rec.get(sf)
        if val and str(val).strip() and str(val).strip() != "None":
            has_existing_sig = True
            break

    if has_existing_sig:
        st.markdown(f"**{T['existing_sig']}:**")
        sig_cols = st.columns(len(s_fields))
        for sf_idx, sig_field in enumerate(s_fields):
            with sig_cols[sf_idx]:
                st.markdown(f"**{sig_field}**")
                val = rec.get(sig_field)
                if val and isinstance(val, list) and len(val) > 0:
                    img_url = val[0].get("signedUrl", val[0].get("url", val[0].get("path", "")))
                    if img_url:
                        if not img_url.startswith("http"):
                            img_url = f"{NOCO_BASE}/{img_url}"
                        st.image(img_url, width=200)
                    else:
                        st.success(T["signed"])
                    # Extract timestamp from filename
                    fname_att = val[0].get("title", val[0].get("filename", ""))
                    import re as _re_sig
                    ts_match = _re_sig.search(r'(\d{8}_\d{6})', fname_att)
                    if ts_match:
                        ts_raw = ts_match.group(1)
                        try:
                            from datetime import datetime as _dts
                            ts_fmt = _dts.strptime(ts_raw, "%Y%m%d_%H%M%S").strftime("%Y-%m-%d %H:%M:%S")
                            st.caption(f"🕐 {ts_fmt}")
                        except Exception:
                            st.caption(f"🕐 {ts_raw}")
                elif val and str(val).strip() and str(val).strip() != "None":
                    st.success(T["signed"])
                else:
                    st.caption(T["not_signed"])
        st.divider()

    # Signature canvases — only for unsigned fields
    if not s_fields:
        st.warning(T["no_sig_field"])
    else:
        # Check if ALL fields already signed
        all_already_signed = all(
            rec.get(sf) and str(rec.get(sf)).strip() not in ("", "None")
            for sf in s_fields
        )
        if all_already_signed:
            st.success(T["sign_confirm_success"])
        else:
            canvas_results = []
            for sf_idx, sig_field in enumerate(s_fields):
                already_signed = (
                    rec.get(sig_field) and
                    str(rec.get(sig_field)).strip() not in ("", "None")
                )
                if already_signed:
                    st.markdown(f"**✅ {sig_field}** — {T['signed']}")
                else:
                    st.markdown(f"**🖊️ {sig_field}**")
                    canvas_result = st_canvas(
                        fill_color="rgba(255, 255, 255, 0)",
                        stroke_width=3,
                        stroke_color="#000000",
                        background_color="#FFFFFF",
                        height=150,
                        width=600,
                        drawing_mode="freedraw",
                        key=f"dlg_canvas_{sf_idx}",
                    )
                    canvas_results.append((sig_field, canvas_result))

            # Single save button
            if st.button(T["sign_submit"], type="primary", use_container_width=True, key="dlg_save"):
                import numpy as np
                from PIL import Image
                from datetime import datetime as _sign_dt

                signed_data = []

                for sig_field, canvas_res in canvas_results:
                    if canvas_res is not None and canvas_res.image_data is not None:
                        img_array = canvas_res.image_data
                        if np.any(img_array[:, :, :3] < 250):
                            img = Image.fromarray(img_array.astype("uint8"), "RGBA")
                            buf = BytesIO()
                            img.save(buf, format="PNG")
                            sig_b64 = base64.b64encode(buf.getvalue()).decode()
                            ts = _sign_dt.now().strftime("%Y%m%d_%H%M%S")
                            signed_data.append((sig_field, sig_b64, ts))

                if not signed_data:
                    st.warning(T["sign_empty"])
                else:
                    all_ok = True
                    for sig_field, sig_b64, ts in signed_data:
                        success, err = noco_upload_signature(
                            tbl_id, rec_id, sig_field, sig_b64,
                            filename=f"{sig_field}_{ts}.png"
                        )
                        if not success:
                            all_ok = False

                    if all_ok:
                        del st.session_state["sign_record"]
                        st.session_state["toast_msg"] = T["sign_confirm_success"]
                        st.session_state["toast_icon"] = "✅"
                        st.rerun()
                    else:
                        st.session_state["toast_msg"] = T["sign_confirm_fail"]
                        st.session_state["toast_icon"] = "❌"
                        st.rerun()


@st.dialog(" ", width="large")
def edit_dialog():
    """Modal dialog for editing a record."""
    edit_data = st.session_state.get("edit_record")
    if not edit_data:
        st.warning("No record selected")
        return
    st.subheader(T["dlg_edit"])

    rec = edit_data["record"]
    rec_id = edit_data["record_id"]
    tbl_id = edit_data["table_id"]
    rec_idx = edit_data["idx"]

    # Compact form style
    st.markdown("""
    <style>
    [data-testid="stDialog"] [data-testid="stVerticalBlock"] {gap: 0.4rem !important;}
    [data-testid="stDialog"] .stTextInput, [data-testid="stDialog"] .stDateInput,
    [data-testid="stDialog"] .stTimeInput {margin-bottom: 0 !important;}
    [data-testid="stDialog"] label p {font-size: 13px !important; margin-bottom: 2px !important;}
    </style>
    """, unsafe_allow_html=True)

    # ── Load form config + NocoDB schema (identical approach to form_dialog) ─
    import re
    from datetime import datetime, date, time

    form_cfg_edit    = load_form_config().get(tbl_id, {})
    field_cfg_edit   = {fc["title"]: fc for fc in form_cfg_edit.get("fields", [])}
    qr_cfg_e         = form_cfg_edit.get("qr_autofill", {})
    qr_delim_e       = qr_cfg_e.get("delimiter", "|")
    qr_map_e         = {m["field"]: m["segment"] for m in qr_cfg_e.get("field_mappings", []) if m.get("field")}

    schema_fields, _ = noco_get_fields(tbl_id)
    SKIP_UIDT_E = {"ID", "CreatedTime", "LastModifiedTime", "CreatedBy",
                   "LastModifiedBy", "Order", "Rollup", "Lookup", "Formula"}
    editable_fields_e = [
        f for f in schema_fields
        if f.get("title", "").lower() not in ("id", "nc_order")
        and not f.get("title", "").lower().startswith("signature")
        and not f.get("system", False)
        and f.get("uidt") not in SKIP_UIDT_E
        and field_cfg_edit.get(f.get("title", ""), {}).get("visible", True)
    ]
    editable_fields_e.sort(
        key=lambda f: field_cfg_edit.get(f.get("title", ""), {}).get("order", f.get("order") or 9999)
    )

    DATE_FMT_MAP_E = {
        "YYYY-MM-DD": "YYYY-MM-DD", "DD/MM/YYYY": "DD/MM/YYYY",
        "MM/DD/YYYY": "MM/DD/YYYY", "DD-MM-YYYY": "DD/MM/YYYY",
        "YYYY/MM/DD": "YYYY/MM/DD",
        "YYYY-MM (วันที่=1)": "YYYY-MM-DD",
        "YYYY (วันที่=01-01)": "YYYY-MM-DD",
    }
    edited_values = {}

    for field in editable_fields_e:
        fname       = field.get("title", "")
        ftype       = field.get("uidt", "SingleLineText")
        cfg_f       = field_cfg_edit.get(fname, {})
        placeholder = cfg_f.get("placeholder", "")
        date_fmt_cfg   = cfg_f.get("date_format", "YYYY-MM-DD") or "YYYY-MM-DD"
        date_input_fmt = DATE_FMT_MAP_E.get(date_fmt_cfg, "YYYY-MM-DD")
        fname_lower    = fname.lower()
        val     = rec.get(fname, "") or ""
        val_str = str(val) if not isinstance(val, (list, dict)) else ""

        # keyword override for SingleLineText
        if ftype == "SingleLineText":
            if any(fname_lower == w or fname_lower.startswith(w+" ") or fname_lower.endswith(" "+w)
                   for w in ["date","日期","วันที่"]):
                ftype = "Date"
            elif any(fname_lower == w or fname_lower.startswith(w+" ") or fname_lower.endswith(" "+w)
                     for w in ["time","时间","เวลา"]):
                ftype = "Time"
        # ── Attachment ────────────────────────────────────────────────────
        if ftype == "Attachment" or (isinstance(val, list) and val and isinstance(val[0], dict)):
            att_list    = val if isinstance(val, list) else []
            remove_idxs = []
            st.markdown(f"**{fname}**")
            if att_list:
                th = '<div style="display:flex;flex-wrap:wrap;gap:8px;margin-bottom:4px">'
                for att in att_list:
                    u = att.get("signedUrl", att.get("url", att.get("path", "")))
                    if not u: continue
                    if not u.startswith("http"): u = f"{NOCO_BASE}/{u.lstrip('/')}"
                    m = att.get("mimetype",""); t = att.get("title", att.get("filename","file"))
                    if m.startswith("image/"):
                        th += f'<a href="{u}" target="_blank"><img src="{u}" style="height:120px;width:auto;border:2px solid #BBDEFB;border-radius:8px;object-fit:cover" /></a>'
                    else:
                        th += f'<a href="{u}" target="_blank" style="display:inline-flex;align-items:center;gap:4px;padding:6px 12px;border:1px solid #BBDEFB;border-radius:6px;font-size:13px;text-decoration:none;color:#1976D2">📎 {t}</a>'
                th += '</div>'
                st.markdown(th, unsafe_allow_html=True)
                for aidx, att in enumerate(att_list):
                    t2 = att.get("title", att.get("filename", f"file_{aidx+1}"))
                    if st.checkbox(f"🗑️ ลบ: {t2}", key=f"edit_rm_{rec_idx}_{fname}_{aidx}"):
                        remove_idxs.append(aidx)
            keep_att  = [a for i, a in enumerate(att_list) if i not in remove_idxs]
            new_files = st.file_uploader(f"+ เพิ่มไฟล์ใหม่ ({fname})",
                type=["jpg","jpeg","png","gif","pdf","xlsx","docx","csv"],
                accept_multiple_files=True, key=f"edit_upload_{rec_idx}_{fname}")
            edited_values[fname] = (keep_att, new_files or [])
            continue

        # ── Date / DateTime ───────────────────────────────────────────────
        if ftype in ("Date", "DateTime"):
            try:
                dv = (datetime.fromisoformat(val_str.replace("Z","+00:00")).date()
                      if "T" in val_str else date.fromisoformat(val_str[:10]))
                edited_values[fname] = st.date_input(fname, value=dv,
                    format=date_input_fmt, key=f"edit_field_{rec_idx}_{fname}")
            except (ValueError, TypeError):
                edited_values[fname] = st.text_input(fname, value=val_str,
                    placeholder=placeholder, key=f"edit_field_{rec_idx}_{fname}")
            continue

        # ── Time ─────────────────────────────────────────────────────────
        if ftype == "Time":
            from st_mui.time_picker import time_picker as _tp
            from datetime import datetime as _now_dt, time as _time
            try:
                if val_str and ":" in val_str:
                    if "T" in val_str:
                        _tv = datetime.fromisoformat(val_str.replace("Z", "+00:00")).time()
                    else:
                        _p = val_str.split(":")
                        _tv = _time(int(_p[0]), int(_p[1]))
                else:
                    _tv = _now_dt.now().time().replace(second=0, microsecond=0)
            except (ValueError, TypeError):
                _tv = _now_dt.now().time().replace(second=0, microsecond=0)
            edited_values[fname] = _tp(label=fname, value=_tv, ampm=False,
                key=f"edit_field_{rec_idx}_{fname}")
            continue

        # ── Number / Decimal ─────────────────────────────────────────────
        if ftype in ("Number", "Decimal", "Currency", "Percent"):
            try: nv = float(val_str) if val_str else 0.0
            except ValueError: nv = 0.0
            edited_values[fname] = st.number_input(fname, value=nv,
                step=1.0 if ftype=="Number" else 0.01,
                key=f"edit_field_{rec_idx}_{fname}")
            continue

        # ── Checkbox ─────────────────────────────────────────────────────
        if ftype == "Checkbox":
            edited_values[fname] = st.checkbox(fname, value=bool(val),
                key=f"edit_field_{rec_idx}_{fname}")
            continue

        # ── LongText ─────────────────────────────────────────────────────
        if ftype == "LongText":
            edited_values[fname] = st.text_area(fname, value=val_str,
                placeholder=placeholder, key=f"edit_field_{rec_idx}_{fname}")
            continue

        # ── SingleSelect ─────────────────────────────────────────────────
        if ftype == "SingleSelect":
            raw = field.get("colOptions") or field.get("options") or {}
            opts_raw = raw.get("options", []) if isinstance(raw, dict) else []
            if isinstance(opts_raw, list):
                opts = [o.get("title", str(o)) if isinstance(o, dict) else str(o) for o in opts_raw]
            else:
                import re as _re_ss
                opts = [o.strip().strip("'") for o in _re_ss.split(r",(?=')", str(opts_raw)) if o.strip()]
            cur_val = str(val).strip() if val else ""
            cur_idx = (opts.index(cur_val) + 1) if cur_val in opts else 0
            edited_values[fname] = st.selectbox(fname, [""] + opts, index=cur_idx,
                key=f"edit_field_{rec_idx}_{fname}")
            continue

        # ── MultiSelect ──────────────────────────────────────────────────
        if ftype == "MultiSelect":
            raw = field.get("colOptions") or field.get("options") or {}
            opts_raw = raw.get("options", []) if isinstance(raw, dict) else []
            if isinstance(opts_raw, list):
                opts = [o.get("title", str(o)) if isinstance(o, dict) else str(o) for o in opts_raw]
            else:
                import re as _re_ms
                opts = [o.strip().strip("'") for o in _re_ms.split(r",(?=')", str(opts_raw)) if o.strip()]
            cur_val = str(val).strip() if val else ""
            cur_sel = [v.strip() for v in cur_val.split(",") if v.strip() in opts]
            sel = st.multiselect(fname, opts, default=cur_sel, key=f"edit_field_{rec_idx}_{fname}")
            edited_values[fname] = ",".join(sel)
            continue

        # ── QR field ─────────────────────────────────────────────────────
        if fname in qr_map_e:
            qr_show_key = f"qr_show_edit_{tbl_id}_{fname}"
            qr_result_key = f"qr_result_edit_{tbl_id}_{fname}"
            tc, cc = st.columns([9,1])
            with tc:
                edited_values[fname] = st.text_input(fname, value=val_str,
                    placeholder=placeholder, key=f"edit_field_{rec_idx}_{fname}")
            with cc:
                st.markdown("<div style='padding-top:28px'>", unsafe_allow_html=True)
                if st.button("📷", key=f"qr_btn_edit_{rec_idx}_{fname}"):
                    st.session_state[qr_show_key] = not st.session_state.get(qr_show_key, False)
                st.markdown("</div>", unsafe_allow_html=True)
            if st.session_state.get(qr_show_key, False):
                st.caption("📷 ส่องกล้องไปที่ QR Code — จะอ่านอัตโนมัติ")
                decoded = qr_webrtc_scanner(
                    widget_key=f"qr_webrtc_edit_{tbl_id}_{fname}",
                    result_key=qr_result_key,
                )
                if decoded:
                    seg = qr_map_e.get(fname, 0)
                    parts = decoded.split(qr_delim_e) if qr_delim_e and seg > 0 else [decoded]
                    fill = decoded if seg == 0 else (parts[seg-1].strip() if seg <= len(parts) else decoded)
                    st.session_state[f"edit_field_{rec_idx}_{fname}"] = fill
                    st.session_state[qr_show_key] = False
                    st.rerun()
            continue

        # ── Default: text input ───────────────────────────────────────────
        edited_values[fname] = st.text_input(fname, value=val_str,
            placeholder=placeholder, key=f"edit_field_{rec_idx}_{fname}")

    st.markdown("")
    if st.button(T["edit_save"], type="primary", use_container_width=True, key="edit_save_btn"):
        changes = {}
        upload_errors = []
        for field in editable_fields_e:
            fname   = field.get("title", "")
            new_val = edited_values.get(fname)
            old_val = rec.get(fname, "") or ""
            if new_val is None:
                continue
            if isinstance(new_val, tuple):
                keep_list, new_files = new_val
                merged = list(keep_list)
                for uf in new_files:
                    att_obj, up_err = noco_upload_attachment(uf)
                    if up_err:
                        upload_errors.append(f"{fname}: {up_err}")
                    elif att_obj:
                        merged.append(att_obj if isinstance(att_obj, dict) else att_obj[0])
                if merged != old_val:
                    changes[fname] = merged
                continue
            new_val_str = str(new_val) if new_val is not None else ""
            old_val_str = str(old_val)
            if new_val_str != old_val_str:
                changes[fname] = new_val_str

        if upload_errors:
            st.error(f"{T['upload_error']}: {', '.join(upload_errors)}")
        elif not changes:
            st.toast("ไม่มีการเปลี่ยนแปลง", icon="ℹ️")
        else:
            success, err = noco_update_record(tbl_id, rec_id, changes)
            if success:
                del st.session_state["edit_record"]
                st.session_state["toast_msg"] = T["delete_success"]
                st.session_state["toast_icon"] = "✅"
                st.cache_data.clear()
                st.rerun()
            else:
                st.session_state["toast_msg"] = f"{T['delete_fail']}: {err}"
                st.session_state["toast_icon"] = "❌"
                st.rerun()


def _generate_pdf(rec, detail_keys, sig_fields, sig_images, rec_idx):
    """Generate PDF bytes using reportlab with CJK font support."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer, Image as RLImage)
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from io import BytesIO as BIO
    import tempfile, os, re

    def strip_thai(text):
        return re.sub(r'[\u0E00-\u0E7F]+', '', str(text)).strip()

    # Register CJK-capable font (Windows system fonts)
    FONT_NAME = "CJKFont"
    font_registered = False
    for font_path in [
        "C:/Windows/Fonts/msyh.ttc",    # Microsoft YaHei
        "C:/Windows/Fonts/simsun.ttc",  # SimSun
        "C:/Windows/Fonts/simhei.ttf",  # SimHei
        "C:/Windows/Fonts/arial.ttf",   # Arial fallback
    ]:
        if os.path.exists(font_path):
            try:
                pdfmetrics.registerFont(TTFont(FONT_NAME, font_path))
                font_registered = True
                break
            except Exception:
                continue
    if not font_registered:
        FONT_NAME = "Helvetica"

    normal = ParagraphStyle("n", fontName=FONT_NAME, fontSize=9, leading=13)
    bold_s = ParagraphStyle("b", fontName=FONT_NAME, fontSize=9, leading=13)
    title_s = ParagraphStyle("t", fontName=FONT_NAME, fontSize=13, leading=18)

    buf = BIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=15*mm, rightMargin=15*mm,
                            topMargin=15*mm, bottomMargin=15*mm)
    story = []
    story.append(Paragraph(f"Record #{rec_idx+1}", title_s))
    story.append(Spacer(1, 6*mm))

    # Detail table
    tdata = []
    for k in detail_keys:
        raw_val = rec.get(k, "") or ""
        # Skip attachment fields — handle separately below
        if isinstance(raw_val, list) and raw_val and isinstance(raw_val[0], dict):
            continue
        val = str(raw_val)
        if not val:
            continue
        tdata.append([Paragraph(strip_thai(k), bold_s), Paragraph(strip_thai(val), normal)])

    if tdata:
        t = Table(tdata, colWidths=[65*mm, None])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#F5F5F5")),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#DDDDDD")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ]))
        story.append(t)
        story.append(Spacer(1, 8*mm))

    # Signatures section
    if sig_fields:
        story.append(Paragraph("Signatures", bold_s))
        story.append(Spacer(1, 4*mm))

        page_w = A4[0] - 30*mm
        n_cols = len(sig_fields)
        col_w = page_w / n_cols

        sig_row_labels = []
        sig_row_imgs = []
        tmp_files = []

        for sf in sig_fields:
            sig_row_labels.append(Paragraph(strip_thai(sf), normal))
            if sf in sig_images:
                try:
                    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".png")
                    tmp.write(sig_images[sf])
                    tmp.close()
                    tmp_files.append(tmp.name)
                    img_w = min(col_w - 8*mm, 55*mm)
                    rl_img = RLImage(tmp.name, width=img_w, height=25*mm)
                    sig_row_imgs.append(rl_img)
                except Exception:
                    sig_row_imgs.append(Paragraph("—", normal))
            else:
                sig_row_imgs.append(Paragraph("—", normal))

        sig_table = Table(
            [sig_row_labels, sig_row_imgs],
            colWidths=[col_w] * n_cols
        )
        sig_table.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#DDDDDD")),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F5F5F5")),
        ]))
        story.append(sig_table)
    else:
        tmp_files = []

    doc.build(story)
    pdf_bytes = buf.getvalue()

    for f in tmp_files:
        try:
            os.unlink(f)
        except Exception:
            pass

    return pdf_bytes


@st.dialog(" ", width="large")
def print_dialog():
    """Generate PDF with record details + signature images."""
    pdata = st.session_state.get("print_record")
    if not pdata:
        return
    st.subheader(T["dlg_print"])

    rec = pdata["record"]
    sig_fields = pdata["sig_fields"]
    rec_idx = pdata["idx"]

    st.markdown(f"**Record #{rec_idx+1}**")

    # Detail preview
    detail_keys = [k for k in rec.keys()
                   if not k.lower().startswith("signature")
                   and k.lower() not in ("nc_order", "createdat", "updatedat")]
    detail_rows = ""
    att_preview = {}
    for i, k in enumerate(detail_keys):
        val = rec.get(k, "") or ""
        if not val:
            continue
        # Attachment — collect separately
        if isinstance(val, list) and val and isinstance(val[0], dict):
            att_preview[k] = val
            continue
        bg = "rgba(0,0,0,0.02)" if i % 2 == 0 else "transparent"
        k_short = (k[:30] + "…") if len(k) > 30 else k
        detail_rows += (
            f'<tr style="background:{bg}">'
            f'<td style="padding:6px 10px;font-weight:600;max-width:160px;width:160px;'
            f'overflow:hidden;text-overflow:ellipsis;white-space:nowrap;'
            f'border-right:1px solid #ddd;color:#555" title="{k}">{k_short}</td>'
            f'<td style="padding:6px 10px;max-width:260px;overflow:hidden;'
            f'text-overflow:ellipsis;white-space:nowrap" title="{val}">{val}</td></tr>'
        )
    st.markdown(
        f'<table style="width:100%;border-collapse:collapse;font-size:12px;border:1px solid #ddd;'
        f'border-radius:6px;overflow:hidden;margin-bottom:12px;table-layout:fixed">'
        f'<tbody>{detail_rows}</tbody></table>',
        unsafe_allow_html=True
    )
    # Show attachment previews
    for att_key, att_list in att_preview.items():
        st.markdown(f"**{att_key}**")
        thumbs_html = '<div style="display:flex;flex-wrap:wrap;gap:8px;margin-bottom:8px">'
        for att in att_list:
            url = att.get("signedUrl", att.get("url", att.get("path", "")))
            if not url:
                continue
            if not url.startswith("http"):
                url = f"{NOCO_BASE}/{url.lstrip('/')}"
            mime = att.get("mimetype", "")
            title = att.get("title", att.get("filename", "file"))
            if mime.startswith("image/"):
                thumbs_html += f'<a href="{url}" target="_blank" title="{title}"><img src="{url}" style="height:100px;width:auto;border:2px solid #BBDEFB;border-radius:8px;object-fit:cover;cursor:zoom-in" /></a>'
            else:
                thumbs_html += f'<a href="{url}" target="_blank" style="display:inline-flex;align-items:center;gap:4px;padding:6px 12px;border:1px solid #BBDEFB;border-radius:6px;font-size:13px;text-decoration:none;color:#1976D2">📎 {title}</a>'
        thumbs_html += '</div>'
        st.markdown(thumbs_html, unsafe_allow_html=True)

    # Helper: extract image bytes from signature field value
    def get_sig_image_bytes(val):
        """NocoDB attachment field can be list or JSON string."""
        import json as _json
        if not val:
            return None
        if isinstance(val, str):
            try:
                val = _json.loads(val)
            except Exception:
                return None
        if isinstance(val, list) and len(val) > 0:
            att = val[0]
            # Try signedPath first (NocoDB v2), then signedUrl, url, path
            img_path = (att.get("signedPath") or att.get("signedUrl") or
                        att.get("url") or att.get("path") or "")
            if img_path:
                if not img_path.startswith("http"):
                    img_path = f"{NOCO_BASE}/{img_path.lstrip('/')}"
                try:
                    r = requests.get(img_path, headers=NOCO_HEADERS, timeout=10)
                    if r.status_code == 200:
                        return r.content
                except Exception:
                    pass
        return None

    # Collect signature images and preview (horizontal layout)
    sig_images = {}
    sig_cols = st.columns(max(len(sig_fields), 1))
    for si, sf in enumerate(sig_fields):
        with sig_cols[si]:
            st.markdown(f"**{sf}**")
            img_bytes = get_sig_image_bytes(rec.get(sf))
            if img_bytes:
                sig_images[sf] = img_bytes
                st.image(img_bytes, use_container_width=True)
            else:
                st.caption(T["not_signed"])

    st.markdown("")

    # Generate PDF directly
    pdf_bytes = _generate_pdf(rec, detail_keys, sig_fields, sig_images, rec_idx)
    st.download_button(
        label="📥 Download PDF",
        data=pdf_bytes,
        file_name=f"record_{rec_idx+1}.pdf",
        mime="application/pdf",
        type="primary",
        use_container_width=True,
        key="download_pdf",
    )


@st.dialog(" ", width="small")
def delete_confirm_dialog():
    """Modal dialog to confirm record deletion."""
    del_data = st.session_state.get("del_record", {})
    record_id = del_data.get("record_id")
    table_id_del = del_data.get("table_id")
    st.warning(T["del_confirm"])
    col_yes, col_no = st.columns(2)
    with col_yes:
        if st.button("✅ ยืนยัน", type="primary", use_container_width=True, key="del_confirm_btn"):
            ok, err = noco_delete_record(table_id_del, record_id)
            if ok:
                st.session_state["toast_msg"] = T["del_success"]
                st.session_state["toast_icon"] = "✅"
                st.rerun()
            else:
                st.error(f"{T['del_fail']}: {err}")
    with col_no:
        if st.button("❌ ยกเลิก", use_container_width=True, key="del_cancel_btn"):
            st.rerun()


@st.dialog("ลบฟอร์ม", width="small")
def form_delete_confirm_dialog():
    """Confirm deletion of a complete report table and its catalog entry."""
    delete_data = st.session_state.get("form_delete_record", {})
    table_id = delete_data.get("table_id")
    report_name = delete_data.get("report_name", table_id)
    if not table_id:
        return

    st.error(f"ต้องการลบฟอร์ม '{report_name}' และข้อมูลทั้งหมดในฟอร์มนี้หรือไม่?")
    st.caption("การลบนี้ไม่สามารถกู้คืนได้")
    yes_col, no_col = st.columns(2)
    with yes_col:
        if st.button("ลบถาวร", type="primary", use_container_width=True, key="form_delete_confirm"):
            table_deleted, table_error = noco_delete_table(table_id)
            if not table_deleted:
                st.error(f"ลบ table ไม่สำเร็จ: {table_error}")
                return
            catalog_deleted, catalog_error = delete_catalog_report(table_id)
            if not catalog_deleted:
                st.error(f"ลบ table สำเร็จ แต่ลบ catalog ไม่สำเร็จ: {catalog_error}")
                return
            st.session_state.pop("form_delete_record", None)
            st.cache_data.clear()
            st.session_state["toast_msg"] = "ลบฟอร์มเรียบร้อยแล้ว"
            st.session_state["toast_icon"] = "✅"
            st.rerun()
    with no_col:
        if st.button("ยกเลิก", use_container_width=True, key="form_delete_cancel"):
            st.session_state.pop("form_delete_record", None)
            st.rerun()


@st.dialog("สร้างแบบฟอร์มใหม่", width="large")
def form_builder_dialog():
    """Admin dialog for creating a NocoDB table and registering its report."""
    if not st.session_state.get("is_admin", False):
        st.error("ต้องเข้าสู่ Admin ก่อน")
        return
    # Keep the dialog open across Streamlit reruns caused by field edits.
    st.session_state["pending_form_builder"] = True

    st.caption("สร้างตารางจริงใน NocoDB แล้วเพิ่มรายการลงหน้า Report อัตโนมัติ")
    dept = st.text_input("แผนก", key="builder_dept")
    report_name = st.text_input("ชื่อรายงาน", key="builder_report_name")

    type_options = {
        "ข้อความสั้น": ("SingleLineText", "varchar", "255"),
        "ข้อความยาว": ("LongText", "text", ""),
        "ตัวเลข": ("Number", "int", ""),
        "ทศนิยม": ("Decimal", "decimal", ""),
        "วันที่": ("Date", "date", ""),
        "Checkbox": ("Checkbox", "tinyint", "1"),
        "เลือกหนึ่งค่า": ("SingleSelect", "varchar", "255"),
        "เลือกหลายค่า": ("MultiSelect", "varchar", "255"),
        "ไฟล์แนบ": ("Attachment", "json", ""),
        "ลายเซ็น": ("Attachment", "json", ""),
    }
    rows_key = "builder_fields"
    if rows_key not in st.session_state:
        st.session_state[rows_key] = [{"name": "", "type": "ข้อความสั้น", "required": False, "options": ""}]

    st.markdown("**Fields**")
    fields = []
    for idx, row in enumerate(st.session_state[rows_key]):
        c_name, c_type, c_req, c_del = st.columns([3, 2, 1, 0.6])
        with c_name:
            name = st.text_input("ชื่อ field", value=row.get("name", ""), key=f"builder_name_{idx}")
        with c_type:
            selected_type = st.selectbox("ชนิดข้อมูล", list(type_options),
                                         index=list(type_options).index(row.get("type", "ข้อความสั้น")),
                                         key=f"builder_type_{idx}")
        with c_req:
            required = st.checkbox("บังคับ", value=row.get("required", False), key=f"builder_req_{idx}")
        with c_del:
            remove = st.button("ลบ", key=f"builder_del_{idx}")
        options = row.get("options", "")
        if selected_type in ("เลือกหนึ่งค่า", "เลือกหลายค่า"):
            options = st.text_input("ตัวเลือกคั่นด้วย comma", value=options, key=f"builder_opts_{idx}")
        if not remove:
            fields.append({"name": name.strip(), "type": selected_type,
                           "required": required, "options": options})

    st.session_state[rows_key] = fields or [{"name": "", "type": "ข้อความสั้น", "required": False, "options": ""}]
    if st.button("➕ เพิ่ม field", key="builder_add_field"):
        st.session_state[rows_key].append({"name": "", "type": "ข้อความสั้น", "required": False, "options": ""})
        st.rerun()

    if st.button("สร้างแบบฟอร์ม", type="primary", use_container_width=True, key="builder_create"):
        import re as _builder_re
        from datetime import datetime as _builder_dt
        clean_fields = [f for f in fields if f["name"]]
        if not dept.strip() or not report_name.strip():
            st.error("กรุณากรอกแผนกและชื่อรายงาน")
            return
        if not clean_fields:
            st.error("กรุณาเพิ่มอย่างน้อย 1 field")
            return
        for field in clean_fields:
            if field["type"] == "ลายเซ็น" and not field["name"].lower().startswith("signature"):
                field["name"] = f"signature {field['name']}"
        names = [f["name"] for f in clean_fields]
        if len(names) != len(set(names)):
            st.error("ชื่อ field ต้องไม่ซ้ำกัน")
            return
        if any(name.lower() in ("id", "nc_order") for name in names):
            st.error("ห้ามใช้ชื่อ field เป็น id หรือ nc_order")
            return

        table_name = "ereport_" + _builder_re.sub(r"[^a-zA-Z0-9_]", "_", report_name.strip()).strip("_").lower()
        table_name = (table_name[:45] + "_" + _builder_dt.now().strftime("%m%d%H%M%S"))[:63]
        noco_columns = []
        config_fields = []
        for order, field in enumerate(clean_fields, start=1):
            uidt, dt, dtxp = type_options[field["type"]]
            is_signature = field["type"] == "ลายเซ็น"
            column = {"column_name": field["name"], "title": field["name"],
                      "uidt": uidt, "dt": dt,
                      "rqd": False if is_signature else field["required"]}
            if dtxp:
                column["dtxp"] = dtxp
            if uidt in ("SingleSelect", "MultiSelect"):
                column["colOptions"] = {"options": [{"title": option.strip()}
                    for option in field["options"].split(",") if option.strip()]}
            noco_columns.append(column)
            config_fields.append({"title": field["name"], "visible": True,
                                  "required": False if is_signature else field["required"], "order": order,
                                  "placeholder": ""})

        noco_cfg = st.secrets.get("nocodb", {})
        base_id = noco_cfg.get("base_id", noco_cfg.get("project_id", ""))
        with st.spinner("กำลังสร้าง table ใน NocoDB..."):
            table_id, create_err = noco_create_table(base_id, report_name.strip(), table_name, noco_columns)
        if create_err or not table_id:
            st.error(f"สร้าง table ไม่สำเร็จ: {create_err or 'ไม่ได้รับ Table_ID จาก NocoDB'}")
            return

        registered, register_err = register_catalog_report(dept.strip(), report_name.strip(), table_id)
        save_form_config({table_id: {"fields": config_fields}})
        st.cache_data.clear()
        st.session_state.pop(rows_key, None)
        if registered:
            st.session_state["toast_msg"] = f"สร้างแบบฟอร์มสำเร็จ: {report_name.strip()}"
            st.session_state["toast_icon"] = "✅"
        else:
            st.session_state["toast_msg"] = f"สร้าง table สำเร็จ แต่ลง catalog ไม่ได้ (Table_ID: {table_id})"
            st.session_state["toast_icon"] = "⚠️"
        st.session_state["builder_register_error"] = register_err
        st.session_state.pop("pending_form_builder", None)
        st.rerun()


@st.dialog(" ", width="large")
def form_config_dialog():
    """Admin dialog to configure field visibility, required, and order per table."""
    cfg_data = st.session_state.get("form_config_record", {})
    table_id_cfg = cfg_data.get("table_id")
    report_name_cfg = cfg_data.get("report_name", "")
    if not table_id_cfg:
        return

    from datetime import time

    st.markdown(f"**{report_name_cfg}**")
    st.caption(T["cfg_dialog_caption"])

    # Fetch fields from NocoDB
    fields, err = noco_get_fields(table_id_cfg)
    if err or not fields:
        st.error(f"{T['cfg_err_schema']}: {err}")
        return

    SKIP_UIDT = {"ID", "CreatedTime", "LastModifiedTime", "CreatedBy", "LastModifiedBy", "Order", "Rollup", "Lookup", "Formula", "Attachment"}
    base_fields = [
        f for f in fields
        if f.get("title", "").lower() not in ("id", "nc_order")
        and not f.get("title", "").lower().startswith("signature")
        and not f.get("system", False)
        and f.get("uidt") not in SKIP_UIDT
    ]
    base_fields.sort(key=lambda f: f.get("order") or 9999)

    # Load existing config
    existing_cfg = load_form_config().get(table_id_cfg, {})
    existing_field_cfg = {fc["title"]: fc for fc in existing_cfg.get("fields", [])}

    DATE_FORMAT_OPTIONS = ["YYYY-MM-DD", "DD/MM/YYYY", "MM/DD/YYYY", "DD-MM-YYYY", "YYYY/MM/DD", "YYYY-MM (วันที่=1)", "YYYY (วันที่=01-01)"]
    # Map display format → actual st.date_input format
    DATE_INPUT_FMT_MAP = {
        "YYYY-MM-DD": "YYYY-MM-DD",
        "DD/MM/YYYY": "DD/MM/YYYY",
        "MM/DD/YYYY": "MM/DD/YYYY",
        "DD-MM-YYYY": "DD/MM/YYYY",   # closest valid
        "YYYY/MM/DD": "YYYY/MM/DD",
        "YYYY-MM (วันที่=1)": "YYYY-MM-DD",
        "YYYY (วันที่=01-01)": "YYYY-MM-DD",
    }
    NUMBER_UIDTS = {"Number", "Decimal", "Currency", "Percent"}

    # Build editable config rows
    new_field_list = []
    st.markdown("---")
    OPS = ["—", ">=", ">", "<=", "<", "="]
    col_h1, col_h2, col_h3, col_h4, col_h5, col_h6, col_h7, col_h8 = st.columns([2, 0.8, 0.8, 0.8, 1.5, 1.5, 1.5, 1.5])
    with col_h1: st.markdown(f"**{T['cfg_col_field']}**")
    with col_h2: st.markdown(f"**{T['cfg_col_visible']}**")
    with col_h3: st.markdown(f"**{T['cfg_col_req']}**")
    with col_h4: st.markdown(f"**{T['cfg_col_order']}**")
    with col_h5: st.markdown(f"**{T['cfg_col_ph']}**")
    with col_h6: st.markdown(f"**{T['cfg_date_format_label']}**")
    with col_h7: st.markdown(f"**{T['cfg_range_min']}**")
    with col_h8: st.markdown(f"**{T['cfg_range_max']}**")

    for idx, field in enumerate(base_fields):
        fname = field.get("title", "")
        ftype = field.get("uidt", "SingleLineText")
        fname_lower = fname.lower()
        is_date_field = (
            ftype in ("Date", "DateTime") or
            any(fname_lower == w or fname_lower.startswith(w + " ") or fname_lower.endswith(" " + w)
                for w in ["date", "日期", "วันที่"])
        )
        is_number_field = ftype in NUMBER_UIDTS
        prev = existing_field_cfg.get(fname, {})
        default_visible  = prev.get("visible", True)
        default_required = prev.get("required", bool(field.get("rqd", 0)))
        default_order    = prev.get("order", field.get("order") or idx + 1)
        default_ph       = prev.get("placeholder", "")
        default_fmt      = prev.get("date_format", "YYYY-MM-DD")
        fmt_idx          = DATE_FORMAT_OPTIONS.index(default_fmt) if default_fmt in DATE_FORMAT_OPTIONS else 0
        default_min_op   = prev.get("range_min_op", "—")
        default_min_val  = prev.get("range_min_val", "")
        default_max_op   = prev.get("range_max_op", "—")
        default_max_val  = prev.get("range_max_val", "")
        min_op_idx = OPS.index(default_min_op) if default_min_op in OPS else 0
        max_op_idx = OPS.index(default_max_op) if default_max_op in OPS else 0

        col_f, col_v, col_r, col_o, col_p, col_fmt, col_mn, col_mx = st.columns([2, 0.8, 0.8, 0.8, 1.5, 1.5, 1.5, 1.5])
        with col_f:
            st.markdown(f"`{fname}`")
        with col_v:
            visible = st.checkbox("visible", value=default_visible, key=f"cfg_vis_{table_id_cfg}_{idx}",
                                  label_visibility="collapsed")
        with col_r:
            required = st.checkbox("required", value=default_required, key=f"cfg_req_{table_id_cfg}_{idx}",
                                   label_visibility="collapsed")
        with col_o:
            order = st.number_input("order", value=default_order, min_value=1, max_value=999,
                                    step=1, key=f"cfg_ord_{table_id_cfg}_{idx}", label_visibility="collapsed")
        with col_p:
            placeholder = st.text_input("placeholder", value=default_ph,
                                        placeholder=T["cfg_ph_hint"],
                                        key=f"cfg_ph_{table_id_cfg}_{idx}",
                                        label_visibility="collapsed")
        with col_fmt:
            if is_date_field:
                date_format = st.selectbox("date_format", DATE_FORMAT_OPTIONS, index=fmt_idx,
                                           key=f"cfg_fmt_{table_id_cfg}_{idx}",
                                           label_visibility="collapsed")
            else:
                date_format = ""
                st.markdown("—")
        with col_mn:
            if is_number_field:
                mn_c1, mn_c2 = st.columns([1, 1])
                with mn_c1:
                    range_min_op = st.selectbox("min_op", OPS, index=min_op_idx,
                                                key=f"cfg_minop_{table_id_cfg}_{idx}",
                                                label_visibility="collapsed")
                with mn_c2:
                    range_min_val = st.text_input("min_val", value=str(default_min_val),
                                                  placeholder="0",
                                                  key=f"cfg_minval_{table_id_cfg}_{idx}",
                                                  label_visibility="collapsed")
            else:
                range_min_op, range_min_val = "—", ""
                st.markdown("—")
        with col_mx:
            if is_number_field:
                mx_c1, mx_c2 = st.columns([1, 1])
                with mx_c1:
                    range_max_op = st.selectbox("max_op", OPS, index=max_op_idx,
                                                key=f"cfg_maxop_{table_id_cfg}_{idx}",
                                                label_visibility="collapsed")
                with mx_c2:
                    range_max_val = st.text_input("max_val", value=str(default_max_val),
                                                  placeholder="100",
                                                  key=f"cfg_maxval_{table_id_cfg}_{idx}",
                                                  label_visibility="collapsed")
            else:
                range_max_op, range_max_val = "—", ""
                st.markdown("—")

        new_field_list.append({
            "title": fname,
            "visible": visible,
            "required": required,
            "order": int(order),
            "placeholder": placeholder,
            "date_format": date_format,
            "range_min_op": range_min_op,
            "range_min_val": range_min_val,
            "range_max_op": range_max_op,
            "range_max_val": range_max_val,
        })
    # ── Date Auto-fill Config ─────────────────────────────────────────────
    st.markdown("---")
    st.markdown(f"**{T['cfg_sec_date_title']}**")

    date_fields = [f.get("title", "") for f in base_fields
                   if f.get("uidt") == "Date" or
                   f.get("uidt") == "DateTime" or
                   f.get("title", "").lower().find("date") >= 0 or
                   f.get("title", "").find("日期") >= 0 or
                   f.get("title", "").find("วันที่") >= 0]

    existing_date_cfg = existing_cfg.get("date_autofill", {})
    none_opt = T["cfg_none"]
    date_opts = [none_opt] + date_fields

    primary_default = existing_date_cfg.get("primary", none_opt)
    if primary_default not in date_opts:
        primary_default = none_opt

    dc1, dc2 = st.columns(2)
    with dc1:
        primary_date = st.selectbox(T["cfg_primary_date"], date_opts,
                                    index=date_opts.index(primary_default),
                                    key=f"cfg_primary_{table_id_cfg}")
    with dc2:
        st.caption("")

    # Secondary date fields — each with its own offset
    existing_secondaries = existing_date_cfg.get("secondaries", [])
    sec_by_field = {s["field"]: s for s in existing_secondaries}

    secondary_fields = [f for f in date_fields if f != primary_date]
    new_secondaries = []

    if secondary_fields and primary_date != none_opt:
        st.markdown(f"_{T['cfg_secondary_dates']}_")
        sh1, sh2, sh3 = st.columns([3, 1.5, 1])
        with sh1: st.markdown(f"**{T['cfg_col_field']}**")
        with sh2: st.markdown(f"**{T['cfg_autofill_enable']}**")
        with sh3: st.markdown(f"**{T['cfg_offset_months']}**")

        for sidx, sfield in enumerate(secondary_fields):
            prev_sec = sec_by_field.get(sfield, {})
            sc1, sc2, sc3 = st.columns([3, 1.5, 1])
            with sc1: st.markdown(f"`{sfield}`")
            with sc2:
                sec_enabled = st.checkbox(T["cfg_autofill_enable"], value=prev_sec.get("enabled", False),
                                          key=f"cfg_sec_en_{table_id_cfg}_{sidx}")
            with sc3:
                sec_offset = st.selectbox("offset", [1, 2, 3, 6, 12],
                                          index=[1,2,3,6,12].index(prev_sec.get("offset_months", 1)) if prev_sec.get("offset_months", 1) in [1,2,3,6,12] else 0,
                                          key=f"cfg_sec_off_{table_id_cfg}_{sidx}",
                                          label_visibility="collapsed",
                                          disabled=not sec_enabled)
            new_secondaries.append({
                "field": sfield,
                "enabled": sec_enabled,
                "offset_months": sec_offset,
            })

    date_autofill_cfg = {
        "primary": primary_date if primary_date != none_opt else "",
        "secondaries": new_secondaries,
    }

    # ── QR Code Auto-fill Config ───────────────────────────────────────────
    st.markdown("---")
    st.markdown(f"**{T['cfg_qr_title']}**")
    st.caption(T["cfg_qr_caption"])

    existing_qr = existing_cfg.get("qr_autofill", {})
    qr_delimiter = st.text_input(T["cfg_qr_delimiter"],
                                 value=existing_qr.get("delimiter", "|"),
                                 key=f"cfg_qr_delim_{table_id_cfg}")

    # List of {field, segment} mappings stored in session_state.
    # We never call st.rerun() here — changes take effect on the next
    # natural interaction (e.g. the user edits another widget) so the
    # dialog stays open. The pending_qr_action key signals what to
    # mutate before rendering the list.
    qr_fields_key = f"cfg_qr_fields_{table_id_cfg}"
    if qr_fields_key not in st.session_state:
        st.session_state[qr_fields_key] = list(existing_qr.get("field_mappings", []))

    # Apply any pending mutation from previous interaction
    pending_qr = st.session_state.pop(f"_pending_qr_{table_id_cfg}", None)
    if pending_qr == "add":
        st.session_state[qr_fields_key].append({"field": "", "segment": 0})
    elif isinstance(pending_qr, int):
        lst = st.session_state[qr_fields_key]
        if 0 <= pending_qr < len(lst):
            lst.pop(pending_qr)

    none_field = T["cfg_none_field"]
    qr_field_options = [none_field] + [f.get("title", "") for f in base_fields]

    if st.session_state[qr_fields_key]:
        qr_h1, qr_h2, qr_h3 = st.columns([3, 2, 1])
        with qr_h1: st.markdown(f"**{T['cfg_qr_col_field']}**")
        with qr_h2: st.markdown(f"**Segment**")
        with qr_h3: st.markdown("")

        for qr_idx, qr_entry in enumerate(st.session_state[qr_fields_key]):
            qc1, qc2, qc3 = st.columns([3, 2, 1])
            with qc1:
                prev_field = qr_entry.get("field", none_field)
                fi = qr_field_options.index(prev_field) if prev_field in qr_field_options else 0
                chosen_field = st.selectbox("qr_field", qr_field_options, index=fi,
                                            key=f"cfg_qr_f_{table_id_cfg}_{qr_idx}",
                                            label_visibility="collapsed")
            with qc2:
                chosen_seg = st.number_input("segment", value=int(qr_entry.get("segment", 0)),
                                             min_value=0, max_value=20, step=1,
                                             key=f"cfg_qr_s_{table_id_cfg}_{qr_idx}",
                                             label_visibility="collapsed")
            with qc3:
                if st.button(T["cfg_qr_remove"], key=f"cfg_qr_del_{table_id_cfg}_{qr_idx}"):
                    st.session_state[f"_pending_qr_{table_id_cfg}"] = qr_idx
            st.session_state[qr_fields_key][qr_idx] = {
                "field": chosen_field if chosen_field != none_field else "",
                "segment": int(chosen_seg),
            }

    if st.button(T["cfg_qr_add_field"], key=f"cfg_qr_add_{table_id_cfg}"):
        st.session_state[f"_pending_qr_{table_id_cfg}"] = "add"

    qr_autofill_cfg = {
        "delimiter": qr_delimiter,
        "field_mappings": [m for m in st.session_state[qr_fields_key] if m.get("field")],
    }

    # ── Generic Webhook Notification Config ────────────────────────────────
    st.markdown("---")
    st.markdown(f"**{T['cfg_notify_title']}**")

    existing_notify = existing_cfg.get("notify", {})
    nc1, nc2 = st.columns([1, 3])
    with nc1:
        notify_enabled = st.checkbox(T["cfg_notify_enable"],
                                     value=existing_notify.get("enabled", False),
                                     key=f"cfg_notify_en_{table_id_cfg}")
    with nc2:
        st.caption("")

    legacy_webhook = existing_notify.get("webhook_url", "")
    notify_webhook = st.text_input(
        T["cfg_notify_webhook"], value=legacy_webhook,
        placeholder="https://example.com/webhook",
        help=T["cfg_notify_webhook_help"],
        key=f"cfg_notify_webhook_{table_id_cfg}")

    DEFAULT_MSG_TEMPLATE = (
        "⚠️ แจ้งเตือนการกรอกข้อมูล\n"
        "รายงาน: {report_name}\n"
        "ยังไม่มีการกรอกข้อมูลใหม่วันนี้ (เริ่มเช็คตั้งแต่ {start_hour})\n"
        "เวลาแจ้งเตือน: {now}"
    )
    notify_template = st.text_area(T["cfg_notify_template"],
                                   value=existing_notify.get("message_template", DEFAULT_MSG_TEMPLATE),
                                   height=120,
                                   key=f"cfg_notify_tpl_{table_id_cfg}")
    st.caption(T["cfg_notify_template_help"])

    st.markdown("---")
    notify_success_enabled = st.checkbox(T["cfg_notify_success_enable"],
                                         value=existing_notify.get("notify_on_success", False),
                                         key=f"cfg_notify_success_en_{table_id_cfg}")
    DEFAULT_SUCCESS_TEMPLATE = (
        "✅ มีการกรอกข้อมูลแล้ว\n"
        "รายงาน: {report_name}\n"
        "เวลาที่ตรวจพบ: {now}"
    )
    notify_success_template = st.text_area(T["cfg_notify_success_template"],
                                           value=existing_notify.get("success_message_template", DEFAULT_SUCCESS_TEMPLATE),
                                           height=100,
                                           disabled=not notify_success_enabled,
                                           key=f"cfg_notify_success_tpl_{table_id_cfg}")

    ntc1, ntc2 = st.columns(2)
    with ntc1:
        start_hour_default = existing_notify.get("start_hour", 7)
        start_minute_default = existing_notify.get("start_minute", 0)
        notify_start_time = st.time_input(T["cfg_notify_start_time"],
                                          value=time(hour=int(start_hour_default), minute=int(start_minute_default)),
                                          key=f"cfg_notify_start_{table_id_cfg}",
                                          step=60)
    with ntc2:
        interval_options = [1, 2, 3, 4, 6, 8, 12]
        interval_default = existing_notify.get("interval_hours", 2)
        interval_idx = interval_options.index(interval_default) if interval_default in interval_options else 1
        notify_interval = st.selectbox(T["cfg_notify_interval"], interval_options,
                                       index=interval_idx,
                                       key=f"cfg_notify_int_{table_id_cfg}")

    if st.button(T["cfg_notify_test"], key=f"cfg_notify_testbtn_{table_id_cfg}"):
        import datetime as _dt_mod
        try:
            test_text = "[TEST] " + notify_template.format(
                report_name=report_name_cfg,
                start_hour=f"{notify_start_time.hour:02d}:{notify_start_time.minute:02d}",
                now=_dt_mod.datetime.now().strftime("%Y-%m-%d %H:%M"),
            )
        except Exception:
            test_text = f"[TEST] {report_name_cfg}\nนี่คือข้อความทดสอบการแจ้งเตือนจากระบบ E Report Board"
        ok_send, err_send = send_webhook_message(notify_webhook, test_text)
        if ok_send:
            st.success(T["cfg_notify_test_ok"])
        else:
            st.error(f"{T['cfg_notify_test_fail']}{err_send}")

    notify_cfg = {
        "enabled": notify_enabled,
        "webhook_url": notify_webhook,
        "message_template": notify_template,
        "notify_on_success": notify_success_enabled,
        "success_message_template": notify_success_template,
        "start_hour": notify_start_time.hour,
        "start_minute": notify_start_time.minute,
        "interval_hours": notify_interval,
        "report_name": report_name_cfg,
    }

    # ── Allow Edit After Submit ────────────────────────────────────────────
    st.markdown("---")
    st.markdown(f"**{T['cfg_allow_edit_title']}**")
    allow_edit_after_submit = st.checkbox(
        T["cfg_allow_edit_label"],
        value=existing_cfg.get("allow_edit_after_submit", False),
        key=f"cfg_allow_edit_{table_id_cfg}",
    )
    st.caption(T["cfg_allow_edit_caption"])

    st.markdown("---")
    if st.button(T["cfg_save"], type="primary", use_container_width=True, key="cfg_save_btn"):
        ok = save_form_config({table_id_cfg: {
            "fields": new_field_list,
            "date_autofill": date_autofill_cfg,
            "qr_autofill": qr_autofill_cfg,
            "notify": notify_cfg,
            "allow_edit_after_submit": allow_edit_after_submit,
        }})
        if ok:
            st.session_state["toast_msg"] = T["cfg_saved"]
            st.session_state["toast_icon"] = "✅"
            st.rerun()
        else:
            st.error(T["cfg_err_schema"])


with tab1:
    # ── Date filter ───────────────────────────────────────────────────────
    from datetime import date as _today_date
    import concurrent.futures

    fc_date1, fc_date2, fc_date3 = st.columns([2, 2, 4])
    with fc_date1:
        sel_date = st.date_input(T["tab1_date_filter"], value=_today_date.today(),
                                 format="YYYY-MM-DD", key="tab1_date")
    with fc_date2:
        sel_dept_t1 = st.selectbox(T["dept"],
                                   [T["all"]] + sorted(df["dept"].dropna().unique().tolist()),
                                   key="tab1_dept")
    with fc_date3:
        search_t1 = st.text_input(T["search"], placeholder=T["search_ph"], key="tab1_search")

    date_str = str(sel_date)

    # Get all tables with Table_ID
    tables_df = df[df["Table_ID"].notna()].copy()
    if sel_dept_t1 != T["all"]:
        tables_df = tables_df[tables_df["dept"] == sel_dept_t1]

    # Parallel fetch records from all tables for selected date
    def fetch_table(row):
        tid = str(row["Table_ID"]).strip()
        recs, err = noco_get_records_by_date(tid, date_str)
        result = []
        for r in recs:
            r["_table_id"] = tid
            r["_report_name"] = row["name"]
            r["_dept"] = row["dept"]
        return recs

    all_records = []
    if sel_date:
        with st.spinner(T["tab1_loading"]):
            table_rows = [row for _, row in tables_df.iterrows()]
            with concurrent.futures.ThreadPoolExecutor(max_workers=10) as ex:
                futures = [ex.submit(fetch_table, row) for row in table_rows]
                for f in concurrent.futures.as_completed(futures):
                    try:
                        all_records.extend(f.result())
                    except Exception:
                        pass

    # Apply search filter
    if search_t1:
        all_records = [r for r in all_records if
                       search_t1.lower() in str(r.get("_report_name", "")).lower() or
                       search_t1.lower() in str(r.get("_dept", "")).lower()]

    if not all_records:
        st.info(T["tab1_no_records"])
    else:
        is_admin = st.session_state.get("is_admin", False)

        # Sort state
        if "t1_sort_col" not in st.session_state:
            st.session_state["t1_sort_col"] = "date"
            st.session_state["t1_sort_asc"] = False

        def sort_icon(col):
            if st.session_state["t1_sort_col"] == col:
                return " ▲" if st.session_state["t1_sort_asc"] else " ▼"
            return " ⇅"

        def set_sort(col):
            if st.session_state["t1_sort_col"] == col:
                st.session_state["t1_sort_asc"] = not st.session_state["t1_sort_asc"]
            else:
                st.session_state["t1_sort_col"] = col
                st.session_state["t1_sort_asc"] = True

        def get_sort_key(rec):
            col = st.session_state["t1_sort_col"]
            if col == "date":
                return str(rec.get("CreatedAt", rec.get("created_at", "")) or "")
            elif col == "name":
                return str(rec.get("_report_name", "")).lower()
            elif col == "recorder":
                return str(rec.get("记录者 Recorder", rec.get("Recorder", "")) or "").lower()
            return ""

        all_records = sorted(all_records, key=get_sort_key,
                             reverse=not st.session_state["t1_sort_asc"])

        st.markdown('<div class="compact-row">', unsafe_allow_html=True)
        st.markdown('<div class="compact-header" style="background:#E3F2FD;border-bottom:2px solid #BBDEFB;border-radius:8px 8px 0 0">', unsafe_allow_html=True)

        # Pre-load form configs for all tables to check allow_edit_after_submit
        _all_form_cfg = load_form_config()

        if is_admin:
            h_no, h_dept, h_rpt, h_date, h_rec, h_status, h_btn, h_edit, h_del, h_print = st.columns([0.4, 1, 1.8, 1, 1.2, 1.8, 0.4, 0.4, 0.4, 0.4])
        else:
            # Check if ANY table in the current result set has allow_edit_after_submit=True
            # so we know whether to reserve an edit column in the header
            _any_user_edit = any(
                _all_form_cfg.get(r.get("_table_id", ""), {}).get("allow_edit_after_submit", False)
                for r in all_records
            )
            if _any_user_edit:
                h_no, h_dept, h_rpt, h_date, h_rec, h_status, h_btn, h_edit, h_print = st.columns([0.4, 1, 1.8, 1, 1.2, 1.8, 0.4, 0.4, 0.4])
            else:
                h_no, h_dept, h_rpt, h_date, h_rec, h_status, h_btn, h_print = st.columns([0.4, 1, 1.8, 1, 1.2, 1.8, 0.4, 0.4])

        with h_no: st.markdown(f"**{T['tbl_no']}**")
        with h_dept: st.markdown(f"**{T['tab1_col_dept']}**")
        with h_rpt:
            if st.button(f"{T['tab1_col_report']}{sort_icon('name')}", key="t1_sort_name_btn", use_container_width=True):
                set_sort("name"); st.rerun()
        with h_date:
            if st.button(f"{T['tbl_date']}{sort_icon('date')}", key="t1_sort_date_btn", use_container_width=True):
                set_sort("date"); st.rerun()
        with h_rec:
            if st.button(f"{T['tbl_recorder']}{sort_icon('recorder')}", key="t1_sort_rec_btn", use_container_width=True):
                set_sort("recorder"); st.rerun()
        with h_status: st.markdown(f"**{T['tbl_sig_status']}**")
        with h_btn: st.markdown(f"**{T['tbl_sign']}**")
        if is_admin:
            with h_edit: st.markdown(f"**{T['tbl_col_edit']}**")
            with h_del:  st.markdown(f"**{T['tbl_col_del']}**")
        elif _any_user_edit:
            with h_edit: st.markdown(f"**{T['tbl_col_edit']}**")
        with h_print: st.markdown(f"**{T['tbl_print']}**")
        st.markdown('</div>', unsafe_allow_html=True)

        for idx, record in enumerate(all_records):
            table_id  = record.get("_table_id", "")
            report_name = record.get("_report_name", "")
            dept_name   = record.get("_dept", "")
            record_id = record.get("Id", record.get("id", idx))
            sig_fields = get_signature_fields([record])

            date_val = record.get("CreatedAt", record.get("created_at", "")) or ""
            if date_val:
                try:
                    import re as _re2
                    from datetime import datetime as _dt, timezone as _tz, timedelta as _td
                    dstr = str(date_val).strip().replace("Z", "+00:00")
                    dstr = _re2.sub(r'(\.\d{6})\d+', r'\1', dstr)
                    dt = _dt.fromisoformat(dstr)
                    if dt.tzinfo is None:
                        dt = dt.replace(tzinfo=_tz.utc)
                    dt_local = dt.astimezone(_tz(_td(hours=7)))
                    date_val = dt_local.strftime("%Y-%m-%d %H:%M")
                except Exception:
                    raw = str(date_val)
                    date_val = raw.replace("T", " ").split("+")[0].split(".")[0][:16]

            recorder = record.get("记录者 Recorder", record.get("Recorder", record.get("recorder", ""))) or ""
            sig_status = ""
            for sf in sig_fields:
                val = record.get(sf)
                if val and str(val).strip() and str(val).strip() != "None":
                    sig_status += f"✅ {sf}  "
                else:
                    sig_status += f"⏳ {sf}  "

            # Check per-record whether this table allows user-edit
            _tbl_cfg = _all_form_cfg.get(table_id, {})
            _allow_user_edit = _tbl_cfg.get("allow_edit_after_submit", False)
            # Determine column layout for this row
            if is_admin:
                c_no, c_dept, c_rpt, c_date, c_rec, c_status, c_btn, c_edit, c_del, c_print = st.columns([0.4, 1, 1.8, 1, 1.2, 1.8, 0.4, 0.4, 0.4, 0.4])
            elif _allow_user_edit or _any_user_edit:
                c_no, c_dept, c_rpt, c_date, c_rec, c_status, c_btn, c_edit, c_print = st.columns([0.4, 1, 1.8, 1, 1.2, 1.8, 0.4, 0.4, 0.4])
            else:
                c_no, c_dept, c_rpt, c_date, c_rec, c_status, c_btn, c_print = st.columns([0.4, 1, 1.8, 1, 1.2, 1.8, 0.4, 0.4])

            with c_no:    st.markdown(f"{idx+1}")
            with c_dept:  st.markdown(dept_name)
            with c_rpt:   st.markdown(report_name)
            with c_date:  st.markdown(date_val)
            with c_rec:   st.markdown(recorder)
            with c_status: st.markdown(sig_status)
            with c_btn:
                if st.button("✍️", key=f"sign_{table_id}_{idx}"):
                    st.session_state["sign_record"] = {
                        "record": record, "record_id": record_id,
                        "table_id": table_id, "sig_fields": sig_fields, "idx": idx,
                    }
                    sign_dialog()
            if is_admin:
                with c_edit:
                    if st.button("✏️", key=f"edit_{table_id}_{idx}"):
                        st.session_state["edit_record"] = {
                            "record": record, "record_id": record_id,
                            "table_id": table_id, "idx": idx,
                        }
                        edit_dialog()
                with c_del:
                    if st.button("🗑️", key=f"del_{table_id}_{idx}", help=T["tbl_col_del"]):
                        st.session_state["del_record"] = {
                            "record_id": record_id, "table_id": table_id, "idx": idx,
                        }
                        st.session_state["pending_dialog"] = "delete_confirm"
            elif _allow_user_edit:
                with c_edit:
                    if st.button("✏️", key=f"edit_{table_id}_{idx}"):
                        st.session_state["edit_record"] = {
                            "record": record, "record_id": record_id,
                            "table_id": table_id, "idx": idx,
                        }
                        edit_dialog()
            elif _any_user_edit:
                # Column exists in layout (reserved for other rows), render empty placeholder
                with c_edit:
                    st.markdown("")
            with c_print:
                all_signed = all(
                    record.get(sf) and str(record.get(sf)).strip() not in ("", "None")
                    for sf in sig_fields
                ) if sig_fields else True
                if all_signed:
                    if st.button("🖨️", key=f"print_{table_id}_{idx}", help=T["print_title"]):
                        st.session_state["print_record"] = {
                            "record": record, "sig_fields": sig_fields, "idx": idx,
                        }
                        print_dialog()
                else:
                    st.markdown("<span title='เซ็นไม่ครบ'>🔒</span>", unsafe_allow_html=True)

            st.markdown('<hr style="margin:2px 0;border-color:#E3F2FD">', unsafe_allow_html=True)

        st.markdown('</div>', unsafe_allow_html=True)

        pending_t1 = st.session_state.pop("pending_dialog", None)
        if pending_t1 == "delete_confirm" and "del_record" in st.session_state:
            delete_confirm_dialog()
# ══════════════════════════════════════════════════════════════════════════
with tab2:
    # ── Filters (main content) ────────────────────────────────────────────
    depts = [T["all"]] + sorted(df["dept"].dropna().unique().tolist())

    fc1, fc2, fc3, fc4 = st.columns([2, 2, 3, 1])
    with fc1:
        selected = st.selectbox(T["dept"], depts, key="tab2_dept")
    with fc2:
        if selected != T["all"]:
            dept_names = sorted(df[df["dept"] == selected]["name"].dropna().unique().tolist())
        else:
            dept_names = sorted(df["name"].dropna().unique().tolist())
        names = [T["all"]] + dept_names
        selected_name = st.selectbox(T["col_name"], names, key="tab2_name")
    with fc3:
        search = st.text_input(T["search"], placeholder=T["search_ph"], key="tab2_search")
    with fc4:
        rows_per_page = st.selectbox(T["rows"], [10, 15, 25, 50], index=1, key="tab2_rows")

    # ── Filter ────────────────────────────────────────────────────────────
    filtered = df.copy()
    if selected != T["all"]:
        filtered = filtered[filtered["dept"] == selected]
    if selected_name != T["all"]:
        filtered = filtered[filtered["name"] == selected_name]
    if search:
        mask = (
            filtered["name"].str.contains(search, case=False, na=False) |
            filtered["dept"].str.contains(search, case=False, na=False)
        )
        filtered = filtered[mask]

    filtered = filtered.reset_index(drop=True)
    total = len(filtered)

    if total == 0:
        st.info(T["no_result"])
        if st.session_state.pop("pending_form_builder", False):
            form_builder_dialog()
    else:
        # ── Pagination state ──────────────────────────────────────────────
        if "tab2_page" not in st.session_state:
            st.session_state["tab2_page"] = 1

        total_pages = max(1, math.ceil(total / rows_per_page))
        # Clamp page without resetting — avoids scroll-to-top on filter change
        if st.session_state["tab2_page"] > total_pages:
            st.session_state["tab2_page"] = total_pages

        start   = (st.session_state["tab2_page"] - 1) * rows_per_page
        page_df = filtered.iloc[start : start + rows_per_page].copy()

        # ── Table ─────────────────────────────────────────────────────────
        def make_link(val, label):
            if pd.notna(val) and str(val).strip().startswith("http"):
                return f'<a href="{str(val).strip()}" target="_blank">{label}</a>'
            return "➖"

        # Header
        is_admin_t2 = st.session_state.get("is_admin", False)
        if is_admin_t2:
            th1, th2, th3, th4, th5, th6, th7, th8, th9, th10 = st.columns([0.5, 1.5, 3, 1, 1, 1, 0.8, 0.6, 0.8, 0.6])
        else:
            th1, th2, th3, th4, th5, th6, th9 = st.columns([0.5, 1.5, 3, 1, 1, 1, 0.8])
        with th1: st.markdown(f"**{T['col_no']}**")
        with th2: st.markdown(f"**{T['col_dept']}**")
        with th3: st.markdown(f"**{T['col_name']}**")
        with th4: st.markdown(f"**{T['col_form']}**")
        with th5: st.markdown(f"**{T['col_table']}**")
        with th6: st.markdown(f"**{T['col_date']}**")
        if is_admin_t2:
            with th7: st.markdown(f"**{T['col_comment']}**")
            with th8: st.markdown(f"**{T['col_cfg']}**")
            with th10: st.markdown("**ลบฟอร์ม**")
        with th9: st.markdown(f"**{T['col_export']}**")
        st.markdown('<hr style="margin:4px 0;border-color:#BBDEFB">', unsafe_allow_html=True)

        comments = load_comments()

        for i, row in page_df.iterrows():
            no      = start + list(page_df.index).index(i) + 1
            dept    = row["dept"] or ""
            name    = row["name"] or ""
            tbl     = make_link(row["Table_link"], T["open"])
            updated = str(row["updated"]) if pd.notna(row["updated"]) else ""
            table_id_row = str(row["Table_ID"]).strip() if pd.notna(row.get("Table_ID")) else None

            if is_admin_t2:
                c1, c2, c3, c4, c5, c6, c7, c8, c9, c10 = st.columns([0.5, 1.5, 3, 1, 1, 1, 0.8, 0.6, 0.8, 0.6])
            else:
                c1, c2, c3, c4, c5, c6, c9 = st.columns([0.5, 1.5, 3, 1, 1, 1, 0.8])
            with c1: st.markdown(f"{no}")
            with c2: st.markdown(f"{dept}")
            with c3: st.markdown(f"{name}")
            with c4:
                if table_id_row:
                    if st.button(T["open"], key=f"form_btn_{i}"):
                        st.session_state["form_record"] = {
                            "table_id": table_id_row,
                            "report_name": name,
                        }
                        st.session_state.pop("pending_form_builder", None)
                        st.session_state["pending_dialog"] = "form"
                else:
                    st.markdown("➖")
            with c5: st.markdown(tbl, unsafe_allow_html=True)
            with c6: st.markdown(f"{updated}")
            if is_admin_t2:
                with c7:
                    if table_id_row:
                        has_comment = bool(comments.get(table_id_row, "").strip())
                        btn_label = "💬✅" if has_comment else "💬"
                        if st.button(btn_label, key=f"comment_btn_{i}", help=T["col_comment"]):
                            st.session_state["comment_record"] = {
                                "table_id": table_id_row,
                                "report_name": name,
                            }
                            st.session_state.pop("pending_form_builder", None)
                            st.session_state["pending_dialog"] = "comment"
                    else:
                        st.markdown("➖")
                with c8:
                    if table_id_row:
                        if st.button("⚙️", key=f"cfg_btn_{i}", help=T["cfg_btn_help"]):
                            st.session_state["form_config_record"] = {
                                "table_id": table_id_row,
                                "report_name": name,
                            }
                            st.session_state.pop("pending_form_builder", None)
                            st.session_state["pending_dialog"] = "form_config"
                    else:
                        st.markdown("➖")
                with c10:
                    if table_id_row:
                        if st.button("🗑️", key=f"form_del_{table_id_row}_{i}", help="ลบฟอร์มทั้งหมด"):
                            st.session_state.pop("pending_form_builder", None)
                            st.session_state.pop("pending_dialog", None)
                            st.session_state["form_delete_record"] = {
                                "table_id": table_id_row,
                                "report_name": name,
                            }
                            st.session_state["pending_form_delete"] = True
                    else:
                        st.markdown("➖")
            with c9:
                if table_id_row:
                    if st.button("📥", key=f"export_btn_{i}", help=T["export_btn_help"]):
                        with st.spinner(T["export_generating"]):
                            all_recs = noco_get_all_records(table_id_row)
                            if not all_recs:
                                st.session_state.pop("pending_export_row", None)
                                st.warning(T["export_no_data"])
                            else:
                                xlsx_bytes, xerr = build_excel_with_images(all_recs, sheet_name=name)
                                if xerr:
                                    st.session_state.pop("pending_export_row", None)
                                    st.error(T["export_error"])
                                else:
                                    safe_name = "".join(c for c in str(name) if c not in '\\/:*?"<>|') or "export"
                                    st.session_state["pending_export_row"] = table_id_row
                                    st.session_state["pending_export_bytes"] = xlsx_bytes
                                    st.session_state["pending_export_filename"] = f"{safe_name}.xlsx"
                                    st.session_state["pending_export_autoclick"] = True

                    if st.session_state.get("pending_export_row") == table_id_row:
                        autoclick = st.session_state.pop("pending_export_autoclick", False)
                        st.download_button(
                            T["export_download"],
                            data=st.session_state["pending_export_bytes"],
                            file_name=st.session_state["pending_export_filename"],
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key=f"export_dl_{i}",
                            on_click=lambda: st.session_state.pop("pending_export_row", None),
                        )
                        if autoclick:
                            import streamlit.components.v1 as components
                            components.html("""
                            <script>
                            setTimeout(function() {
                                const btns = window.parent.document.querySelectorAll('[data-testid="stDownloadButton"] button');
                                if (btns.length > 0) { btns[btns.length - 1].click(); }
                            }, 250);
                            </script>
                            """, height=0)
                else:
                    st.markdown("➖")
            st.markdown('<hr style="margin:2px 0;border-color:#E3F2FD">', unsafe_allow_html=True)

        # Open dialog AFTER loop — prevents rerun on filter interaction
        form_delete_pending = st.session_state.pop("pending_form_delete", False)
        builder_pending = st.session_state.pop("pending_form_builder", False)
        if form_delete_pending:
            form_delete_confirm_dialog()
        elif builder_pending:
            form_builder_dialog()
        else:
            pending = st.session_state.pop("pending_dialog", None)
            if pending == "form" and "form_record" in st.session_state:
                form_dialog()
            elif pending == "comment" and "comment_record" in st.session_state:
                comment_dialog()
            elif pending == "form_config" and "form_config_record" in st.session_state:
                form_config_dialog()

        # ── Pagination ────────────────────────────────────────────────────
        st.divider()
        pc1, pc2, pc3, pc4, pc5 = st.columns([2, 1, 1, 1, 2])

        with pc1:
            st.caption(T["page_info"](st.session_state["tab2_page"], total_pages, start+1, min(start+rows_per_page, total), total))
        with pc2:
            if st.button("«", disabled=st.session_state["tab2_page"] == 1, key="pg_first"):
                st.session_state["tab2_page"] = 1; st.rerun()
        with pc3:
            if st.button("‹", disabled=st.session_state["tab2_page"] == 1, key="pg_prev"):
                st.session_state["tab2_page"] -= 1; st.rerun()
        with pc4:
            if st.button("›", disabled=st.session_state["tab2_page"] >= total_pages, key="pg_next"):
                st.session_state["tab2_page"] += 1; st.rerun()
        with pc5:
            if st.button("»", disabled=st.session_state["tab2_page"] >= total_pages, key="pg_last"):
                st.session_state["tab2_page"] = total_pages; st.rerun()
